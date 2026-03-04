"""
Excel to JSON Policy Schema Converter

Parses Guardian Policy Schema Excel files and converts them to JSON format
following the GuardianPolicySchemaWithDefinitions model.

Usage:
    python excel_to_json.py <input.xlsx> [output.json]

Example:
    python excel_to_json.py Test_Root_Schema_MultiStep_UsingBuilder.xlsx schema_output.json
"""

import json
import logging
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from ..models.guardian_policy_schema import (
    EnumOptions,
    GuardianPolicySchema,
    HelpTextStyle,
    MetadataBase,
    SchemaField,
    SchemaReference,
    VisibilityCondition,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelFormulaParser:
    """Parse Excel formulas back to visibility conditions."""

    @staticmethod
    def parse_visibility_formula(formula: str, field_key_to_answer_col: dict[str, str]) -> dict:
        """
        Parse Excel visibility formula (like =NOT(EXACT(C5,"Type-A"))) back to condition dict.

        Args:
            formula: Excel formula string (with or without leading =)
            field_key_to_answer_col: Mapping from field keys to their Answer column cells (e.g., {"type": "C5"})

        Returns:
            Dictionary representing the condition structure
        """
        if not formula:
            return {}

        # Remove leading '=' if present
        formula = formula.strip()
        if formula.startswith("="):
            formula = formula[1:]

        # Check for NOT wrapper (invert)
        invert = False
        if formula.startswith("NOT(") and formula.endswith(")"):
            invert = True
            formula = formula[4:-1]  # Remove NOT( and )

        # Parse the expression
        condition = ExcelFormulaParser._parse_expression(formula, field_key_to_answer_col)

        if invert:
            return {"invert": True, "condition": condition}
        return {"invert": False, "condition": condition}

    @staticmethod
    def _parse_expression(expr: str, field_key_to_answer_col: dict[str, str]) -> dict:
        """Parse boolean expression recursively."""
        expr = expr.strip()

        # Try to match logical operators (AND, OR)
        for op in ["AND", "OR"]:
            if expr.startswith(f"{op}("):
                # Find matching parentheses
                parts = ExcelFormulaParser._split_function_args(expr[len(op) + 1 : -1])
                if len(parts) == 2:
                    return {
                        "operator": op,
                        "left": ExcelFormulaParser._parse_expression(
                            parts[0], field_key_to_answer_col
                        ),
                        "right": ExcelFormulaParser._parse_expression(
                            parts[1], field_key_to_answer_col
                        ),
                    }

        # Try to match comparison operators (EXACT)
        if expr.startswith("EXACT("):
            parts = ExcelFormulaParser._split_function_args(expr[6:-1])
            if len(parts) == 2:
                left = ExcelFormulaParser._parse_operand(parts[0], field_key_to_answer_col)
                right = ExcelFormulaParser._parse_operand(parts[1], field_key_to_answer_col)
                return {"operator": "EQUAL", "left": left, "right": right}

        raise ValueError(f"Unable to parse expression: {expr}")

    @staticmethod
    def _split_function_args(args: str) -> list[str]:
        """Split function arguments by comma, respecting nested parentheses."""
        parts = []
        current = []
        depth = 0

        for char in args:
            if char == "," and depth == 0:
                parts.append("".join(current).strip())
                current = []
            else:
                if char == "(":
                    depth += 1
                elif char == ")":
                    depth -= 1
                current.append(char)

        if current:
            parts.append("".join(current).strip())

        return parts

    @staticmethod
    def _parse_operand(operand: str, field_key_to_answer_col: dict[str, str]) -> dict:
        """Parse an operand (can be cell reference or constant value)."""
        operand = operand.strip()

        # Check if it's a cell reference (like C5, AA10, etc.)
        if re.match(r"^[A-Z]+\d+$", operand):
            # Find which field key this cell reference belongs to
            field_key = None
            for key, cell_ref in field_key_to_answer_col.items():
                if cell_ref == operand:
                    field_key = key
                    break

            if field_key:
                return {"field_key_ref": field_key}
            logger.warning(f"Cell reference {operand} not found in field mapping")
            return {"field_key_ref": operand}

        # Otherwise it's a constant value
        # Remove quotes if present
        if operand.startswith('"') and operand.endswith('"'):
            operand = operand[1:-1]

        return {"value": operand}


class ExcelPolicySchemaParser:
    """Parse Excel policy schema files back to JSON."""

    # Expected column names (must match mapper.py)
    EXPECTED_COLUMNS = [
        "Required Field",
        "Field Type",
        "Parameter",
        "Visibility",
        "Question",
        "Allow Multiple Answers",
        "Answer",
        "Default",
        "Suggest",
        "Key",
    ]

    # Expected metadata keys
    EXPECTED_METADATA = ["Description", "Schema Type"]

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = load_workbook(excel_path, data_only=False)  # Keep formulas
        self.schemas: list[GuardianPolicySchema] = []
        # Map from enum sheet identifier (field key) to list of options
        self.sheet_name_to_options: dict[str, list[str]] = {}

    def parse(self) -> list[GuardianPolicySchema]:
        """Parse the entire Excel workbook."""
        logger.info(f"Parsing Excel file: {self.excel_path}")

        # First pass: Parse enum sheets (to extract options)
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.endswith(" (enum)"):
                self._parse_enum_sheet(self.workbook[sheet_name], sheet_name)

        # Second pass: Parse schema sheets
        for sheet_name in self.workbook.sheetnames:
            if not sheet_name.endswith(" (enum)") and not sheet_name.endswith(" (tool)"):
                schema = self._parse_schema_sheet(self.workbook[sheet_name], sheet_name)
                if schema:
                    self.schemas.append(schema)

        logger.info(
            f"Parsed {len(self.schemas)} schemas and {len(self.sheet_name_to_options)} enums"
        )

        return self.schemas

    def _parse_enum_sheet(self, sheet: Worksheet, sheet_name: str):
        """Parse an enum definition sheet and store options."""
        logger.info(f"Parsing enum sheet: {sheet_name}")

        # Row 1: Schema name (not needed)
        # Row 2: Field name (not needed)
        # Row 3+: Enum options

        options = []
        row_idx = 3
        while True:
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value is None:
                break
            options.append(str(cell_value))
            row_idx += 1

        # Extract identifier from sheet name (remove " (enum)" suffix)
        # This is the field key that was used to create the enum sheet
        enum_key = sheet_name[:-7]  # Remove " (enum)"

        # Store options by sheet name (both with and without suffix for flexibility)
        self.sheet_name_to_options[enum_key] = options
        self.sheet_name_to_options[sheet_name] = options

        logger.info(f"  Found enum '{enum_key}' with {len(options)} options")

    def _parse_schema_sheet(self, sheet: Worksheet, sheet_name: str) -> GuardianPolicySchema | None:
        """Parse a schema sheet."""
        logger.info(f"Parsing schema sheet: {sheet_name}")

        # Row 1: Worksheet name (title)
        # schema_name = sheet.cell(row=1, column=1).value
        schema_name = sheet.title
        if not schema_name:
            logger.warning(f"  Sheet {sheet_name} has no schema name, skipping")
            return None

        # Rows 2-3: Metadata (Description, Schema Type)
        metadata = self._parse_metadata(sheet)

        # Find table header row (should be row 4)
        header_row_idx = 4
        columns = self._parse_header_row(sheet, header_row_idx)

        if not columns:
            logger.warning(f"  Sheet {sheet_name} has no valid table headers, skipping")
            return None

        # Parse data rows starting from row 5
        fields, field_key_to_answer_col = self._parse_data_rows(sheet, header_row_idx + 1, columns)

        # Second pass: Parse visibility conditions now that we have field_key_to_answer_col
        self._parse_visibility_conditions(
            sheet, header_row_idx + 1, columns, fields, field_key_to_answer_col
        )

        logger.info(f"  Parsed schema '{schema_name}' with {len(fields)} fields")

        return GuardianPolicySchema(
            schema_name=schema_name, metadata=metadata, schema_fields=fields
        )

    def _parse_metadata(self, sheet: Worksheet) -> MetadataBase:
        """Parse metadata rows (rows 2-3)."""
        metadata_dict = {}

        for row_idx in [2, 3]:
            key = sheet.cell(row=row_idx, column=1).value
            value = sheet.cell(row=row_idx, column=2).value

            if key:
                metadata_dict[key] = value or ""

        return MetadataBase(
            description=metadata_dict.get("Description", ""),
            schema_type=metadata_dict.get("Schema Type", "Verifiable Credentials"),
        )

    def _parse_header_row(self, sheet: Worksheet, row_idx: int) -> dict[str, int]:
        """Parse header row and return column name to index mapping."""
        columns = {}

        for col_idx in range(1, 20):  # Check up to 20 columns
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                columns[str(cell_value)] = col_idx

        return columns

    def _parse_data_rows(
        self, sheet: Worksheet, start_row: int, columns: dict[str, int]
    ) -> tuple[list[SchemaField], dict[str, str]]:
        """
        Parse data rows and extract fields, respecting outline levels.
        Only parse fields at outline_level 0 (top-level fields for this schema).
        Nested fields belong to sub-schemas and will be parsed separately.

        Returns:
            Tuple of (fields list, field_key_to_answer_col mapping)
        """
        fields = []
        field_key_to_answer_col = {}
        print(f"Parsing data rows starting at row {start_row} Sheet: {sheet.title!a}")
        row_idx = start_row
        while True:
            # Check if this row is empty (no Question or Key)
            question_col = columns.get("Question")
            key_col = columns.get("Key")

            if not question_col or not key_col:
                print("Missing Question or Key column in header")
                break

            question_value = sheet.cell(row=row_idx, column=question_col).value
            key_value = sheet.cell(row=row_idx, column=key_col).value

            # Stop if both question and key are empty
            if not question_value and not key_value:
                break

            # Check outline level - only parse top-level fields (outline_level == 0)
            outline_level = sheet.row_dimensions[row_idx].outline_level

            if outline_level == 0:
                # Parse this row as a field
                field = self._parse_field_row(sheet, row_idx, columns)
                if field:
                    fields.append(field)

                    # Track field key to Answer column mapping for visibility parsing
                    answer_col = columns.get("Answer")
                    if answer_col and field.key:
                        from openpyxl.utils import get_column_letter

                        col_letter = get_column_letter(answer_col)
                        field_key_to_answer_col[field.key] = f"{col_letter}{row_idx}"

                    # If this field references a sub-schema, skip its nested rows
                    if isinstance(field.field_type, SchemaReference):
                        # Skip rows until we return to outline_level 0
                        row_idx += 1
                        while row_idx < sheet.max_row:
                            next_outline = sheet.row_dimensions[row_idx].outline_level
                            if next_outline == 0:
                                break
                            row_idx += 1
                        continue

            row_idx += 1

        return fields, field_key_to_answer_col

    def _parse_field_row(
        self, sheet: Worksheet, row_idx: int, columns: dict[str, int]
    ) -> SchemaField | None:
        """Parse a single field row."""
        # Extract all column values
        row_data = {}
        for col_name, col_idx in columns.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data[col_name] = self._extract_cell_value(cell, col_name)

        # Required fields
        key = row_data.get("Key")
        if not key:
            return None

        # Parse field_type (can be string or SchemaReference)
        field_type_value = row_data.get("Field Type", "String")
        field_type = self._parse_field_type(field_type_value)

        # Parse parameter (can be string, EnumOptions, or HelpTextStyle)
        parameter_value = row_data.get("Parameter", "")
        parameter = self._parse_parameter(parameter_value, field_type)

        # Visibility will be parsed in second pass
        visibility = ""

        return SchemaField(
            required_field=row_data.get("Required Field", "No"),
            field_type=field_type,
            parameter=parameter,
            visibility=visibility,  # Will be filled in second pass
            question=row_data.get("Question", ""),
            allow_multiple_answers=row_data.get("Allow Multiple Answers", "No"),
            answer=row_data.get("Answer", ""),
            key=key,
            suggest=row_data.get("Suggest", ""),
            default=row_data.get("Default", ""),
        )

    def _parse_visibility_conditions(
        self,
        sheet: Worksheet,
        start_row: int,
        columns: dict[str, int],
        fields: list[SchemaField],
        field_key_to_answer_col: dict[str, str],
    ):
        """Second pass: Parse visibility conditions for all fields."""
        visibility_col = columns.get("Visibility")
        if not visibility_col:
            return

        for field_idx, field in enumerate(fields):
            row_idx = start_row + field_idx
            cell = sheet.cell(row=row_idx, column=visibility_col)

            # Check if cell has a formula
            if cell.data_type == "f" and cell.value:
                formula = str(cell.value)
                try:
                    condition_dict = ExcelFormulaParser.parse_visibility_formula(
                        formula, field_key_to_answer_col
                    )
                    if condition_dict:
                        # Convert dict to VisibilityCondition model
                        field.visibility = VisibilityCondition(**condition_dict)
                except Exception as e:
                    logger.warning(
                        f"Failed to parse visibility formula for field '{field.key}': {formula}. Error: {e}"
                    )
                    field.visibility = ""
            elif cell.value:
                # Non-formula value
                if str(cell.value).upper() == "HIDDEN":
                    field.visibility = "Hidden"
                elif str(cell.value).upper() in ["TRUE", "FALSE"]:
                    # Static TRUE/FALSE
                    field.visibility = ""
                else:
                    field.visibility = ""

    def _extract_cell_value(self, cell: Cell, col_name: str) -> Any:
        """Extract value from cell, handling hyperlinks and special cases."""
        if cell.hyperlink:
            # This is a hyperlink (likely a sheet reference)
            # Return the display text
            return str(cell.value) if cell.value else ""

        # For regular cells, return the value as string (Guardian schema expects strings)
        if cell.value is None:
            return ""

        # Convert to string for consistency
        return str(cell.value)

    def _parse_field_type(self, value: Any) -> str | SchemaReference:
        """Parse field type, detecting schema references."""
        if not value:
            return "String"

        value_str = str(value).strip()

        # List of known primitive field types
        KNOWN_FIELD_TYPES = {
            "Enum",
            "Number",
            "Integer",
            "String",
            "Pattern",
            "Boolean",
            "Date",
            "Time",
            "DateTime",
            "Duration",
            "URL",
            "URI",
            "Email",
            "Image",
            "Help Text",
            "GeoJSON",
            "Prefix",
            "Postfix",
            "HederaAccount",
            "Auto-Calculate",
            "File",
        }

        # If it's a known field type, return as-is
        if value_str in KNOWN_FIELD_TYPES:
            return value_str

        # Otherwise, check if there's a corresponding sheet (schema reference)
        if value_str in self.workbook.sheetnames:
            # This is a schema reference
            return SchemaReference(unique_schema_name_ref=value_str)

        # If not in known types and not a sheet, log warning and assume schema reference
        logger.warning(f"Unknown field type '{value_str}' - assuming sub-schema reference")

        # Assume it's a schema reference to a sheet that will be parsed
        return SchemaReference(unique_schema_name_ref=value_str)

    def _parse_parameter(
        self, value: Any, field_type: str | SchemaReference
    ) -> str | EnumOptions | HelpTextStyle:
        """Parse parameter field based on field type."""
        if not value:
            if field_type == "Enum":
                # Enum fields must have EnumOptions, return empty options if not found
                return EnumOptions(options=[], unique_name="")
            return ""

        value_str = str(value).strip()

        # If field_type is Enum, parameter should be EnumOptions
        if field_type == "Enum" and value_str:
            # Look up enum options from the parsed enum sheets
            # value_str is typically the sheet name (field key)
            enum_key = value_str
            if f"{value_str} (enum)" in self.workbook.sheetnames:
                enum_key = value_str

            # Get options from the stored enum data
            options = self.sheet_name_to_options.get(enum_key, [])
            if not options:
                # Try with (enum) suffix
                options = self.sheet_name_to_options.get(f"{enum_key} (enum)", [])

            return EnumOptions(options=options, unique_name=enum_key)

        # If field_type is Help Text, parameter should be HelpTextStyle
        if field_type == "Help Text":
            # Try to parse as JSON
            try:
                if value_str.startswith("{"):
                    style_dict = json.loads(value_str)
                    return HelpTextStyle(**style_dict)
            except Exception:
                pass
            # Return default HelpTextStyle
            return HelpTextStyle()

        # For Prefix/Postfix, parameter is the symbol
        if field_type in ("Prefix", "Postfix"):
            return value_str

        # For Pattern, parameter is the regex
        if field_type == "Pattern":
            return value_str

        # Default: return as string
        return value_str


def main():
    """Main entry point for CLI usage."""
    if len(sys.argv) < 2:
        print("Usage: python excel_to_json.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    # Auto-generate output path if not provided
    if not output_path:
        input_file = Path(input_path)
        output_path = input_file.with_suffix(".json")

    # Parse Excel file
    parser = ExcelPolicySchemaParser(input_path)
    schemas = parser.parse()

    # Convert to JSON - save as array of schemas
    json_output = [schema.model_dump(mode="json", exclude_none=True) for schema in schemas]

    # Write to file
    output_file = Path(output_path)
    output_file.write_text(json.dumps(json_output, indent=2), encoding="utf-8")

    logger.info(f"Successfully converted {input_path} to {output_path}")
    print(f"Output written to: {output_path}")


if __name__ == "__main__":
    main()
