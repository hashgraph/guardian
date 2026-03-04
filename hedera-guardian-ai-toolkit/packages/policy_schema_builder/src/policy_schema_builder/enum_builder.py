import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

# class MetadataParams(BaseModel):
#     """
#     Parameters for metadata key-value pairs.
#     """
#     key: str = Field(
#         description="The metadata key name."
#     )
#     value: Optional[str] = Field(
#         "",
#         description="The metadata value."
#     )


class ExcelPolicyEnumBuilderParams(BaseModel):
    sheet_name: str = Field(description="The name of the Excel sheet for the enum.")
    schema_name: str = Field(description="The name of the schema containing the enum.")
    field_name: str = Field(
        description="The name of the field for which the enum is defined. (ussualy equals to question text)"
    )
    options: list[str]


class ExcelPolicyEnumBuilder:
    """
    Builder for Excel enum definitions in policy schema based on enum defined in:
    https://github.com/hashgraph/guardian/blob/main/docs/methodology-digitization-handbook/_shared/artifacts/schema-template-excel.xlsx
    """

    def __init__(self, sheet: Worksheet, params: ExcelPolicyEnumBuilderParams):
        self.options_start_row = 0

        self.sheet = sheet
        self.params = params

    def _setup_column_widths(self):
        """Set column widths to match the original Excel file."""
        self.sheet.column_dimensions["A"].width = 30.0
        self.sheet.column_dimensions["B"].width = 50.0

    def _apply_header_title(self, cell):
        """
        Apply header title styling:
        - Bold font
        - 14pt size
        - Borders on all sides (thin, black)
        """
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.border = Border(
            left=Side(style="thin", color="FF000000"),
            right=Side(style="thin", color="FF000000"),
            top=Side(style="thin", color="FF000000"),
            bottom=Side(style="thin", color="FF000000"),
        )

    def _apply_value_cell(self, cell):
        """
        Apply value cell styling:
        - Normal font (11pt)
        - Wrap text enabled
        - Borders on all sides (thin, black)
        """
        cell.font = Font(name="Calibri", size=11, bold=False)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="FF000000"),
            right=Side(style="thin", color="FF000000"),
            top=Side(style="thin", color="FF000000"),
            bottom=Side(style="thin", color="FF000000"),
        )

    def _apply_option_cell(self, cell):
        """
        Apply option cell styling:
        - Normal font (11pt)
        - Wrap text enabled
        - Borders on left, right, and top only (no bottom border within merged cells)
        """
        cell.font = Font(name="Calibri", size=11, bold=False)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="FF000000"),
            right=Side(style="thin", color="FF000000"),
        )

    def _set_schema_name(self):
        """Set the schema name row with proper styling."""
        # Header title cell
        cell_title = self.sheet.cell(row=1, column=1, value="Schema name")
        self._apply_header_title(cell_title)

        # Value cell
        cell_value = self.sheet.cell(row=1, column=2, value=self.params.schema_name)
        self._apply_value_cell(cell_value)

        return 1

    def _set_field_name(self, start_row: int):
        """Set the field name row with proper styling."""
        # Header title cell
        cell_title = self.sheet.cell(row=start_row, column=1, value="Field name")
        self._apply_header_title(cell_title)

        # Value cell
        cell_value = self.sheet.cell(row=start_row, column=2, value=self.params.field_name)
        self._apply_value_cell(cell_value)

        return 1

    def _set_enum_options(self, start_row: int):
        """
        Set enum options with proper styling and merging.
        Each option is merged across columns A and B.
        """
        for index, option in enumerate(self.params.options):
            current_row = start_row + index

            # Set the cell value
            cell_value = self.sheet.cell(row=current_row, column=1, value=option)
            self._apply_option_cell(cell_value)

            # Merge cells A and B for this row
            self.sheet.merge_cells(
                start_row=current_row, end_row=current_row, start_column=1, end_column=2
            )

            # Apply styling to the merged cell (column B also needs styling)
            cell_b = self.sheet.cell(row=current_row, column=2)
            self._apply_option_cell(cell_b)

        return len(self.params.options)

    def build(self):
        """
        Build the complete Excel enum sheet with all styling applied.
        Returns the styled worksheet.
        """

        self._setup_column_widths()

        current_row = 1
        current_row += self._set_schema_name()
        current_row += self._set_field_name(current_row)
        self.options_start_row = current_row
        current_row += self._set_enum_options(current_row)

        return self.sheet

    def get_options_start_position(self) -> str:
        """Get the starting row index for enum options."""
        return f"A{self.options_start_row}"


if __name__ == "__main__":
    # Example usage
    wb = Workbook()
    enum_params = ExcelPolicyEnumBuilderParams(
        schema_name="Employee Onboarding",
        field_name="Employment Type",
        options=["Full-time", "Part-time", "Contractor", "Intern"],
    )

    sheet = wb.create_sheet(title="EmploymentTypeEnum")

    enum_builder = ExcelPolicyEnumBuilder(sheet=sheet, params=enum_params)
    enum_builder.build()
    wb.save("policy_schema_with_enum.xlsx")
