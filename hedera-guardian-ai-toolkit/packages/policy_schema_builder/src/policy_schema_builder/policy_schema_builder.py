import logging
from typing import Union

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from .models import ExcelPolicySchemaParams, FieldColumnParams, MetadataParams, WorksheetNameParams

logger = logging.getLogger(__name__)


def try_to_digitize(value):
    # Strip whitespace if it's a string
    val_str = str(value).strip()

    try:
        # Check if it's an integer (like "10" or "-10")
        if float(val_str).is_integer():
            return int(float(val_str))

        # Check if it's a float (like "10.4" or "-10.5")
        return float(val_str)

    except ValueError:
        # If conversion fails, it's a regular string (like "abc")
        return val_str


class ExcelPolicySchemaBuilder:
    """
    Builder class to create Excel files following the policy schema template.
    Includes all styling, validations, and formatting from the original template.
    Now accepts ExcelPolicySchemaParams for flexible configuration.
    """

    # Color definitions from template
    HEADER_FILL_COLOR = "FFD8E4BC"  # Light green for title and metadata headers
    TABLE_HEADER_FILL_COLOR = "FFFABF8F"  # Orange for table headers
    DATA_ROW_FILL_COLOR = "FFFDE9D9"  # Light pink for data rows (outline level 0)
    NESTED_ROW_FILL_COLOR = "FFF2F2F2"  # Light gray for nested rows (outline level 1+)
    NESTED_ROW_HEADER_FILL_COLOR = "E2E2E2"
    CELL_FILL_WHITE = "FFFFFFFF"  # White for metadata values

    # Font definitions
    HEADER_FONT_SIZE = 14.0
    DATA_FONT_SIZE = 11.0

    # Default column width when not specified
    DEFAULT_COLUMN_WIDTH = 15.0

    # Border style
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Border without top (for data rows continuation)
    BORDER_NO_TOP = Border(
        left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin")
    )

    # Border for nested rows (dashed left/right, solid bottom, no top)
    NESTED_BORDER_NO_TOP = Border(
        left=Side(style="dashed"), right=Side(style="dashed"), bottom=Side(style="thin")
    )

    # Row height
    STANDARD_ROW_HEIGHT = 18.0

    def __init__(self, params: ExcelPolicySchemaParams, sheet: Worksheet):
        """
        Initialize the builder with ExcelPolicySchemaParams.

        Args:
            params: ExcelPolicySchemaParams containing all configuration
        """

        ExcelPolicySchemaParams.model_validate(params)

        self._params = params
        self._num_columns = len(params.table_columns)
        self._sheet = sheet
        self._setup_column_widths()

    def _setup_column_widths(self):
        """Set up column widths based on table_columns params."""
        active_sheet = self._sheet
        for col_index, column in enumerate(self._params.table_columns, start=1):
            col_letter = self.get_column_letter(col_index)
            width = column.width if column.width is not None else self.DEFAULT_COLUMN_WIDTH
            active_sheet.column_dimensions[col_letter].width = width

    @staticmethod
    def get_column_letter(col_index: int) -> str:
        """Convert column index (1-based) to Excel column letter."""
        result = ""
        while col_index > 0:
            col_index, remainder = divmod(col_index - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _apply_header_style(
        self,
        cell,
        fill_color: str,
        font_size: float = None,
        bold: bool = True,
        center: bool = False,
    ):
        """Apply header styling to a cell."""
        if font_size is None:
            font_size = self.HEADER_FONT_SIZE
        cell.font = Font(name="Calibri", size=font_size, bold=bold)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.border = self.THIN_BORDER
        if center:
            cell.alignment = Alignment(horizontal="center")

    def _apply_metadata_style(self, cell, fill_color: str = None):
        """Apply styling to metadata value cells (always thin borders all around)."""
        if fill_color is None:
            fill_color = self.CELL_FILL_WHITE
        cell.font = Font(name="Calibri", size=self.DATA_FONT_SIZE)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.border = self.THIN_BORDER

    def _apply_table_data_style(
        self, cell, is_first_row: bool = False, is_nested: bool = False, is_header_row: bool = False
    ):
        """
        Apply styling to table data cells.

        Args:
            cell: The cell to style
            fill_color: Fill color (defaults to DATA_ROW_FILL_COLOR or NESTED_ROW_FILL_COLOR)
            is_first_row: True if this is the first data row (borders header, needs top border)
            is_nested: True if this is a nested row (uses dashed borders)
        """

        fill_color = self.DATA_ROW_FILL_COLOR

        # First row after header gets thin border all around
        if is_first_row:
            cell.border = self.THIN_BORDER
        # Nested rows get dashed left/right, thin bottom, no top
        elif is_nested:
            cell.border = self.NESTED_BORDER_NO_TOP
            fill_color = (
                self.NESTED_ROW_FILL_COLOR
                if not is_header_row
                else self.NESTED_ROW_HEADER_FILL_COLOR
            )
        # Regular rows get thin left/right/bottom, no top
        else:
            cell.border = self.BORDER_NO_TOP

        cell.font = Font(name="Calibri", size=self.DATA_FONT_SIZE)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    def _set_row_height(self, row_num: int, height: float = None):
        """Set row height."""
        if height is None:
            height = self.STANDARD_ROW_HEIGHT
        self._sheet.row_dimensions[row_num].height = height

    def _set_worksheet_name(self, start_row: int = 1) -> int:
        """
        Set the worksheet name and create the title row.
        Merges cells across all columns and applies styling.

        Args:
            start_row: The row to start at (default 1)

        Returns:
            The next available row number
        """
        active_sheet = self._sheet
        name = self._params.schema_name

        # Set the title in cell A1
        title_cell = active_sheet.cell(row=start_row, column=1, value=name)

        # Merge cells across all columns
        end_col = self._num_columns
        active_sheet.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=end_col
        )

        # Apply styling to the merged title cell
        self._apply_header_style(
            title_cell,
            self.HEADER_FILL_COLOR,
            font_size=self.HEADER_FONT_SIZE,
            bold=True,
            center=True,
        )

        # Set row height
        self._set_row_height(start_row)

        return start_row + 1

    def _set_metadata(self, start_row: int = 2) -> int:
        """
        Set metadata key-value pairs with styling and validations.

        Args:
            start_row: The row to start at (default 2)

        Returns:
            The next available row number
        """
        active_sheet = self._sheet
        end_col = self._num_columns

        for metadata_item in self._params.metadata:
            # Set the key in column A with header styling
            key_cell = active_sheet.cell(row=start_row, column=1, value=metadata_item.key)
            self._apply_header_style(key_cell, self.HEADER_FILL_COLOR)

            # Set the value in column B
            value_cell = active_sheet.cell(row=start_row, column=2, value=metadata_item.value)
            self._apply_metadata_style(value_cell, fill_color=self.CELL_FILL_WHITE)

            # Merge cells B through end_col for the value
            active_sheet.merge_cells(
                start_row=start_row, start_column=2, end_row=start_row, end_column=end_col
            )

            # Apply border to all merged cells
            for col in range(2, end_col + 1):
                cell = active_sheet.cell(row=start_row, column=col)
                cell.border = self.THIN_BORDER

            # Add data validation if value_restrictions are specified
            if metadata_item.value_restrictions:
                validation_values = ",".join(metadata_item.value_restrictions)
                validation_range = f"B{start_row}:{self.get_column_letter(end_col)}{start_row}"
                dv = DataValidation(
                    type="list",
                    formula1=f'"{validation_values}"',
                    showDropDown=False,
                    allowBlank=True,
                )
                active_sheet.add_data_validation(dv)
                dv.add(validation_range)

            # Set row height
            self._set_row_height(start_row)

            start_row += 1

        return start_row

    def _set_fields_headers(self, start_row: int) -> int:
        """
        Set the table headers row with styling.

        Args:
            start_row: The row to place headers

        Returns:
            The next available row number
        """
        active_sheet = self._sheet

        for col_index, column in enumerate(self._params.table_columns, start=1):
            header_cell = active_sheet.cell(row=start_row, column=col_index, value=column.value)
            self._apply_header_style(
                header_cell,
                self.TABLE_HEADER_FILL_COLOR,
                font_size=self.HEADER_FONT_SIZE,
                bold=True,
            )

        # Set row height
        self._set_row_height(start_row)

        return start_row + 1

    def _set_fields_validations(self, start_row: int, end_row: int):
        """
        Set data validations for columns that have value_restrictions.

        Args:
            start_row: First data row
            end_row: Last data row
        """
        active_sheet = self._sheet

        for col_index, column in enumerate(self._params.table_columns, start=1):
            if column.value_restrictions:
                validation_values = ",".join(column.value_restrictions)
                col_letter = self.get_column_letter(col_index)

                dv = DataValidation(
                    type="list",
                    formula1=f'"{validation_values}"',
                    showDropDown=False,
                    allowBlank=True,
                )
                active_sheet.add_data_validation(dv)
                dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    def _set_fields_values(self, start_row: int) -> int:
        """
        Set the data values with styling and validations.
        Supports passing callable functions that receive the cell object
        and return the value to set.

        Now supports nested rows (lists within lists) for creating expandable/collapsible groups.
        Also supports Row and SubRows instances for structured row definitions.

        Args:
            start_row: The row to start placing data

        Returns:
            The next available row number
        """
        rows = self._params.table_rows

        # Convert Row and SubRows instances to list representation
        converted_rows = self._convert_rows_to_lists(rows)

        # Process rows recursively to handle nested structures
        # First row flag is tracked to apply proper border styling
        return self._process_rows_recursive(
            converted_rows, start_row, outline_level=0, first_data_row=start_row
        )

    def _convert_rows_to_lists(self, rows: list) -> list:
        """
        Convert Row and SubRows instances to their list representations.
        This allows the recursive processing logic to work with both
        structured instances and raw lists.

        Args:
            rows: List of rows (can be lists, Row instances, or SubRows instances)

        Returns:
            List of converted rows as nested lists
        """
        converted = []
        for row in rows:
            if isinstance(row, self.Row):
                converted.append(row.to_list())
            elif isinstance(row, self.SubRows):
                converted.append(row.to_nested_list())
            else:
                # Raw list - keep as is (for backward compatibility)
                converted.append(row)
        return converted

    def _process_rows_recursive(
        self, rows: list, current_row: int, outline_level: int, first_data_row: int
    ) -> int:
        """
        Recursively process rows, handling nested structures for expandable/collapsible groups.

        Args:
            rows: List of rows to process (may contain nested lists)
            current_row: The current row number to write to
            outline_level: The outline/grouping level (0 = top level, 1+ = nested)
            first_data_row: The first data row number (for border styling)

        Returns:
            The next available row number after processing all rows
        """
        for row in rows:
            # Check if this is a nested structure (list of lists)
            if self._is_nested_row(row):
                # This is a parent row with children
                # First row is the heading (stays at current outline level)
                # Subsequent rows are the subrows (get increased outline level)
                for idx, subrow in enumerate(row):
                    if idx == 0:
                        # Heading row - keep at current outline level
                        is_first_row = current_row == first_data_row
                        self._write_data_row(
                            subrow, current_row, outline_level, is_first_row, is_header_row=True
                        )
                        current_row += 1
                    # Subrows - process with increased outline level
                    elif self._is_nested_row(subrow):
                        # Nested subrow
                        current_row = self._process_rows_recursive(
                            [subrow], current_row, outline_level + 1, first_data_row
                        )
                    else:
                        # Regular subrow
                        is_first_row = current_row == first_data_row
                        self._write_data_row(subrow, current_row, outline_level + 1, is_first_row)
                        current_row += 1
            else:
                # This is a regular data row
                is_first_row = current_row == first_data_row
                self._write_data_row(row, current_row, outline_level, is_first_row)
                current_row += 1

        return current_row

    def _is_nested_row(self, row: list) -> bool:
        """
        Check if a row is a nested structure (contains sub-rows).
        A row is considered nested if all its elements are lists.

        Args:
            row: The row to check

        Returns:
            True if the row is nested, False otherwise
        """
        if not isinstance(row, list):
            return False

        # Check if all elements are lists (indicating nested structure)
        return all(isinstance(item, list) for item in row)

    def _write_data_row(
        self,
        row: list,
        row_num: int,
        outline_level: int,
        is_first_row: bool,
        is_header_row: bool = False,
    ):
        """
        Write a single data row to the worksheet with proper styling and outline level.
        Applies different fill colors and borders based on the outline level:
        - Level 0: Light pink fill, thin left/right/bottom borders (no top), except first row
        - Level 1+: Light gray fill, dashed left/right borders, thin bottom (no top)
        - First row: Always gets thin border all around (including top)

        Args:
            row: The row data to write
            row_num: The row number to write to
            outline_level: The outline/grouping level for this row
            is_first_row: True if this is the first data row after the header
        """
        active_sheet = self._sheet

        # Determine if this is a nested row
        is_nested = outline_level > 0

        for col_index, cell_value in enumerate(row, start=1):
            data_cell = active_sheet.cell(row=row_num, column=col_index)

            # Apply table data style with appropriate parameters
            self._apply_table_data_style(
                data_cell,
                is_first_row=is_first_row,
                is_nested=is_nested,
                is_header_row=is_header_row,
            )

            # Check if cell_value is callable (a function)
            if callable(cell_value):
                # Call the function with the cell object
                cell_value(data_cell, active_sheet)
            else:
                # Set the value directly
                data_cell.value = cell_value
                # Prevent openpyxl from interpreting text that starts with '='
                # as a formula (e.g. "= Value of model input variable …").
                # Actual formulas are injected via FormulaCell callables above.
                if isinstance(cell_value, str) and cell_value.startswith("="):
                    data_cell.data_type = "s"

        # Apply outline level for grouping (Excel's expandable/collapsible feature)
        if outline_level > 0:
            active_sheet.row_dimensions[row_num].outline_level = outline_level
            # Configure summary rows to appear above detail rows (instead of below)
            active_sheet.sheet_properties.outlinePr.summaryBelow = False

    def _set_fields_table(self, start_row: int) -> int:
        """
        Set the complete fields table with headers and data.

        Args:
            start_row: The row to start the table

        Returns:
            The next available row number
        """
        if not self._params.table_columns:
            return start_row

        start_row = self._set_fields_headers(start_row)
        first_data_row = start_row
        end_row = self._set_fields_values(start_row)

        # Apply data validations to all data rows
        if end_row > first_data_row:
            self._set_fields_validations(first_data_row, end_row - 1)

        return end_row

    def build(self) -> "ExcelPolicySchemaBuilder":
        """
        Build the complete Excel file with all sections.
        Call this method to populate the workbook with schema data.

        Returns:
            self for method chaining
        """
        start_row = 1

        start_row = self._set_worksheet_name(start_row)
        start_row = self._set_metadata(start_row)
        start_row = self._set_fields_table(start_row)

        return self

    class CellBase:
        def __init__(self, title: str):
            self.title = try_to_digitize(title)
            self.value_restrictions = []
            self.number_format = None

        def __call__(self, cell: Cell, worksheet: Worksheet):
            cell.value = self.title
            # Prevent text that starts with '=' from being treated as a formula.
            if isinstance(self.title, str) and self.title.startswith("="):
                cell.data_type = "s"

            if self.value_restrictions:
                validation_values = ",".join(self.value_restrictions)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{validation_values}"',
                    showDropDown=False,
                    allowBlank=True,
                )
                worksheet.add_data_validation(dv)
                dv.add(cell.coordinate)

            if self.number_format:
                cell.number_format = self.number_format

        def set_value_restrictions(self, restrictions: list[str]):
            self.value_restrictions = restrictions
            return self

        def set_number_format(self, number_format: str):
            self.number_format = number_format
            return self

        def set_prefix_format(self, prefix: str):
            """
            Apply prefix number format (e.g., "$"#,##0.00 for currency).

            Args:
                prefix: The prefix string to display before the number

            Returns:
                self for method chaining
            """
            escaped = prefix.replace('"', '""')
            self.number_format = f'"{escaped}"#,##0.00'
            return self

        def set_postfix_format(self, postfix: str):
            """
            Apply postfix number format (e.g., #,##0.00"kg" for units).

            Args:
                postfix: The postfix string to display after the number

            Returns:
                self for method chaining
            """
            escaped = postfix.replace('"', '""')
            self.number_format = f'#,##0.00"{escaped}"'
            return self

    class SheetRefCell(CellBase):
        """
        Callable class to create enum cell with hyperlink sheet.
        """

        def __init__(self, title: str, sheet_name: str, focus_row: int = 1):
            super().__init__(title)
            self.sheet_name = sheet_name
            self.focus_row = focus_row

        def __call__(self, cell: Cell, worksheet: Worksheet):
            super().__call__(cell, worksheet)

            current_font = cell.font
            cell.font = Font(
                name=current_font.name,
                size=current_font.size,
                bold=current_font.bold,
                italic=current_font.italic,
                underline="single",  # Add underline
                color=colors.Color(theme=10),  # Hyperlink blue
            )

            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                location=f"'{self.sheet_name}'!A{self.focus_row}",
            )

    class FormulaCell(CellBase):
        """
        Callable class to create a formula cell.
        """

        def __init__(self, formula: str):
            super().__init__(title="")
            self.formula = formula

        def __call__(self, cell: Cell, worksheet: Worksheet) -> str:  # noqa: ARG002
            cell.value = f"={self.formula}"

    class Row:
        """
        Represents a single row with cell values.
        This class wraps a regular data row to make it explicit and consistent.

        Example:
            Row(["Yes", "String", "", "", "What is your name?", "No", "John", "name"])
        """

        def __init__(self, *values):
            """
            Initialize a Row with cell values.

            Args:
                *values: Variable number of cell values for the row
            """
            self.values = list(values)

        def to_list(self) -> list:
            """Convert Row to list representation."""
            return self.values

        def add_cell(self, value):
            """Add a cell value to the row."""
            self.values.append(value)

    class SubRows:
        """
        Represents nested sub-rows with a heading row and child rows.
        This creates an expandable/collapsible group in Excel.

        The heading_row appears at the parent level, and subrows are nested underneath it.

        Example:
            SubRows(
                heading_row=Row("Yes", "Object", "", "", "Address", "No", "", "address"),
                subrows=[
                    Row("Yes", "String", "", "", "Street", "No", "Main St", "street"),
                    Row("Yes", "String", "", "", "City", "No", "NYC", "city"),
                ]
            )
        """

        def __init__(
            self,
            heading_row: "ExcelPolicySchemaBuilder.Row",
            subrows: list[
                Union["ExcelPolicySchemaBuilder.Row", "ExcelPolicySchemaBuilder.SubRows"]
            ],
            is_nested=False,
        ):
            """
            Initialize SubRows with a heading row and child rows.

            Args:
                heading_row: The parent/heading row (Row instance)
                subrows: List of child rows (can be Row or SubRows instances for deeper nesting)
            """
            self.heading_row = heading_row
            self.subrows = subrows
            self.is_nested = is_nested

        def to_nested_list(self) -> list:
            """
            Convert SubRows to nested list representation.
            Returns a list where the first element is the heading row,
            followed by converted child rows.
            """
            result = []

            heading_cells = self.heading_row.to_list()
            heading_cells_modified = []

            if self.is_nested:
                for cell in heading_cells:
                    heading_cells_modified.append(cell)

                result.append(heading_cells_modified)
            else:
                result.append(heading_cells)

            for subrow in self.subrows:
                if isinstance(subrow, ExcelPolicySchemaBuilder.Row):
                    result.append(subrow.to_list())
                elif isinstance(subrow, ExcelPolicySchemaBuilder.SubRows):
                    result.append(subrow.to_nested_list())
                else:
                    # Fallback for backward compatibility with raw lists
                    result.append(subrow)
            return result


# Example usage
if __name__ == "__main__":
    # EXAMPLE USAGE AND TESTING OF THE BUILDER

    worksheet_name = "Test Schema Name"
    params = ExcelPolicySchemaParams(
        worksheet_name=WorksheetNameParams(value=worksheet_name),
        schema_name=worksheet_name,
        metadata=[
            MetadataParams(key="Description", value="Test schema for generating Excel from JSON"),
            MetadataParams(
                key="Schema Type",
                value="Verifiable Credentials",
                value_restrictions=[
                    "Verifiable Credentials",
                    "Encrypted Verifiable Credential",
                    "Sub-Schema",
                ],
            ),
        ],
        table_columns=[
            FieldColumnParams(value="Required Field", value_restrictions=["Yes", "No"], width=20.0),
            FieldColumnParams(value="Field Type", width=40.0),
            FieldColumnParams(value="Parameter", width=20.0),
            FieldColumnParams(value="Visibility", width=15.0),
            FieldColumnParams(value="Question", width=70.0),
            FieldColumnParams(
                value="Allow Multiple Answers", value_restrictions=["Yes", "No"], width=30.0
            ),
            FieldColumnParams(value="Answer", width=50.0),
            FieldColumnParams(value="Key", width=15.0),
        ],
        table_rows=[
            # Simple row using Row class
            ExcelPolicySchemaBuilder.Row(
                "Yes",
                "String",
                "",
                "",
                "What is your full name?",
                "No",
                "John Doe",
                "full_name",
            ),
            # Row with SheetRefCell
            ExcelPolicySchemaBuilder.Row(
                "Yes",
                "String",
                "",
                "",
                "What is your email address?",
                "No",
                "john.doe@example.com",
                "email_address",
            ),
            # Complex nested structure using SubRows
            ExcelPolicySchemaBuilder.SubRows(
                heading_row=ExcelPolicySchemaBuilder.Row(
                    "Yes",
                    ExcelPolicySchemaBuilder.SheetRefCell(
                        title="Sub Schema 1",
                        sheet_name="SubSchema1Sheet",
                    ),
                    "",
                    "",
                    "What is your sub full name?",
                    "No",
                    "Sub John Doe",
                    "sub_full_name",
                ),
                subrows=[
                    ExcelPolicySchemaBuilder.Row(
                        "No",
                        "Enum",
                        ExcelPolicySchemaBuilder.SheetRefCell(
                            title="Sub Enum", sheet_name="SubEnumSheet"
                        ),
                        "",
                        "What is your sub country?",
                        "No",
                        "Sub USA",
                        "sub_country",
                    ),
                    ExcelPolicySchemaBuilder.Row(
                        "Yes",
                        "String",
                        "",
                        "",
                        "What is your email address?",
                        "No",
                        "john.doe@example.com",
                        "email_address",
                    ),
                    # Deeply nested SubRows
                    ExcelPolicySchemaBuilder.SubRows(
                        heading_row=ExcelPolicySchemaBuilder.Row(
                            "Yes",
                            ExcelPolicySchemaBuilder.SheetRefCell(
                                title="Sub Sub Schema", sheet_name="SubSubSchemaSheet2"
                            ),
                            "",
                            "",
                            "What is your sub sub full name?",
                            "No",
                            "Sub Sub John Doe",
                            "sub_sub_full_name",
                        ),
                        is_nested=True,
                        subrows=[
                            ExcelPolicySchemaBuilder.Row(
                                "No",
                                "Enum",
                                ExcelPolicySchemaBuilder.SheetRefCell(
                                    title="Sub Sub Enum",
                                    sheet_name="SubSubEnumSheet",
                                ),
                                "",
                                "What is your sub sub country?",
                                "No",
                                "Sub Sub USA",
                                "sub_sub_country",
                            ),
                        ],
                    ),
                ],
            ),
            # Another simple row with Enum
            ExcelPolicySchemaBuilder.Row(
                "No",
                "Enum",
                ExcelPolicySchemaBuilder.SheetRefCell(
                    title="CountryEnum",
                    sheet_name="CountryEnumSheet",
                ),
                "",
                "What is your country?",
                "No",
                ExcelPolicySchemaBuilder.CellBase("USA").set_value_restrictions(
                    ["USA", "Canada", "UK", "Australia"]
                ),
                "country",
            ),
        ],
    )

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    builder = ExcelPolicySchemaBuilder(params=params, sheet=ws)
    builder.build()
    wb.save("test_policy_schema.xlsx")

    print("Excel file created successfully!")
