import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

from .enum_builder import ExcelPolicyEnumBuilder, ExcelPolicyEnumBuilderParams
from .models import ExcelPolicySchemaParams
from .models.guardian_policy_schema import SchemaType
from .policy_schema_builder import (
    ExcelPolicySchemaBuilder,
)

logger = logging.getLogger(__name__)


class CompoundExcelBuilderParams(BaseModel):
    excel_policy_schema_params: list[ExcelPolicySchemaParams] = []
    enum_params: list[ExcelPolicyEnumBuilderParams] = []
    worksheet_file_name: str = Field(
        description="Can containe .xlsx extension or not, it will be added if missing.",
    )


class CompoundExcelBuilder:
    def __init__(
        self,
        compound_params: CompoundExcelBuilderParams,
    ):
        self.excel_policy_schema_params = compound_params.excel_policy_schema_params
        self.enum_params = compound_params.enum_params.copy()
        self.worksheet_file_name = self._add_xlsx_extension_if_missing(
            compound_params.worksheet_file_name
        )

        # Validate that at least one root schema exists when creating a new workbook
        # if not self.existing_excel_file_path and self.excel_policy_schema_params:
        #     has_root = any(
        #         self._is_root_policy_schema(schema) for schema in self.excel_policy_schema_params
        #     )
        #     if not has_root:
        #         raise ValueError(
        #             "At least one root schema is required when creating a new workbook. "
        #             "Root schemas must have 'Schema Type' metadata set to 'Verifiable Credentials' "
        #             "or 'Encrypted Verifiable Credential'."
        #         )

        self.workbook: Workbook
        self.sheet: Worksheet

    def _check_excel_file_exists(self, file_path: str) -> bool:
        path = Path(file_path)
        return path.is_file()

    def _add_xlsx_extension_if_missing(self, file_name: str) -> str:
        if not file_name.lower().endswith(".xlsx"):
            file_name += ".xlsx"
        return file_name

    def _is_root_policy_schema(self, schema_param: ExcelPolicySchemaParams) -> bool:
        metadata = schema_param.metadata or []
        result = next(
            (meta_param for meta_param in metadata if meta_param.key == "Schema Type"),
            None,
        )

        if not result:
            logger.error("Schema Type metadata not found, cannot determine if root schema.")
            return False

        return result.value not in [SchemaType.TOOL_INTEGRATION, SchemaType.SUB_SCHEMA]

    def _get_root_policy_schema_params(self) -> ExcelPolicySchemaParams | None:
        for schema_param in self.excel_policy_schema_params:
            if self._is_root_policy_schema(schema_param):
                return schema_param
        return None

    def _create_new_sheet(self, sheet_name: str) -> Worksheet:
        """
        Create or replace a sheet with the given name.
        Sheet name should already be unique and properly formatted by the mapper.

        Args:
            sheet_name: The unique sheet name (pre-resolved by mapper)

        Returns:
            The created or existing worksheet
        """
        if sheet_name in self.workbook.sheetnames:
            logger.warning(
                f"Sheet with name '{sheet_name}' already exists in the workbook, existing sheet will be replaced!"
            )
            return self.workbook[sheet_name]

        return self.workbook.create_sheet(title=sheet_name)

    def create_enum_sheet(self, enum_param: ExcelPolicyEnumBuilderParams):
        enum_sheet = self._create_new_sheet(enum_param.sheet_name)
        enum_builder = ExcelPolicyEnumBuilder(enum_sheet, enum_param)
        enum_builder.build()

    def init_main_policy_schema_sheet(self):
        # Creates empty workbook
        self.workbook = Workbook()
        self.sheet = self.workbook.active

        root_schema = self._get_root_policy_schema_params()

        if root_schema:
            # Fills first sheet with root schema
            main_schema_builder = ExcelPolicySchemaBuilder(root_schema, self.sheet)
            main_schema_builder.build()
            self.sheet.title = root_schema.worksheet_name.value

            # enums
            for enum_param in self.enum_params:
                if enum_param.schema_name == root_schema.worksheet_name.value:
                    logger.debug(f"Creating enum sheet '{enum_param.sheet_name}' for root schema.")
                    self.create_enum_sheet(enum_param)

            self.enum_params = [
                e for e in self.enum_params if e.schema_name != root_schema.worksheet_name.value
            ]

        else:
            logger.warning(
                "No root schema found among the provided schema parameters, the first sheet will be empty."
            )

        rest_schemas = [
            schema for schema in self.excel_policy_schema_params if schema != root_schema
        ]

        # Creates sheets for the rest of the schemas after first sheet
        for schema_param in rest_schemas:
            schema_sheet = self._create_new_sheet(schema_param.worksheet_name.value)
            schema_builder = ExcelPolicySchemaBuilder(schema_param, schema_sheet)
            schema_builder.build()
            # enums
            for enum_param in self.enum_params:
                if enum_param.schema_name == schema_param.worksheet_name.value:
                    logger.debug(
                        f"Creating enum sheet '{enum_param.sheet_name}' for schema '{schema_param.worksheet_name.value}'."
                    )
                    self.create_enum_sheet(enum_param)

            self.enum_params = [
                e for e in self.enum_params if e.schema_name != schema_param.worksheet_name.value
            ]

        # Orphaned enums
        for enum_param in self.enum_params:
            logger.debug(
                f"Creating orphaned enum sheet '{enum_param.sheet_name}' not linked to any schema."
            )
            self.create_enum_sheet(enum_param)

        self.enum_params = []

    def build(self):
        self.init_main_policy_schema_sheet()

    def save(self, output_dir=""):
        """
        Save the filled workbook to the specified path.

        Args:
            output_dir (str): Directory where the filled Excel file will be saved
        """
        # title = sanitize_filename(self.worksheet_file_name)

        # exact_short_date = datetime.now().strftime("%Y%m%d_%H%M%S")

        path_dir = Path(output_dir)
        path_dir.mkdir(parents=True, exist_ok=True)

        full_path = path_dir / self.worksheet_file_name

        logger.info(f"Saving Excel file to: {full_path}")

        self.workbook.save(full_path)

        return full_path


# Example usage
if __name__ == "__main__":
    # EXAMPLE USAGE AND TESTING OF THE BUILDER
    from .models import (
        ExcelPolicySchemaParams,
        FieldColumnParams,
        MetadataParams,
        WorksheetNameParams,
    )

    # SINGLE STEP CREATION

    question_for_subschema = "Provide additional user info"

    sub_schema_worksheet_name = "Sub Additional User Info12345678"
    sub_policy_schema_params = ExcelPolicySchemaParams(
        worksheet_name=WorksheetNameParams(value=sub_schema_worksheet_name),
        schema_name=sub_schema_worksheet_name,
        metadata=[
            MetadataParams(key="Description", value=question_for_subschema),
            MetadataParams(
                key="Schema Type",
                value="Sub-Schema",
                value_restrictions=[
                    "Verifiable Credentials",
                    "Encrypted Verifiable Credential",
                    "Sub-Schema",
                ],
            ),
        ],
        table_columns=[
            FieldColumnParams(value="Required Field", value_restrictions=["Yes", "No"], width=20.0),
            FieldColumnParams(value="Field Type", width=40.0),
            FieldColumnParams(value="Parameter", width=20.0),
            FieldColumnParams(value="Visibility", width=15.0),
            FieldColumnParams(value="Question", width=70.0),
            FieldColumnParams(
                value="Allow Multiple Answers", value_restrictions=["Yes", "No"], width=30.0
            ),
            FieldColumnParams(value="Answer", width=50.0),
            FieldColumnParams(value="Key", width=15.0),
        ],
        table_rows=[["No", "Integer", "", "FALSE", "What is your age?", "No", "30", "age"]],
    )

    question_for_second_subschema = "Provide employment details"

    second_sub_schema_worksheet_name = "Sub Employment Details324234234"
    second_sub_policy_schema_params = ExcelPolicySchemaParams(
        worksheet_name=WorksheetNameParams(value=second_sub_schema_worksheet_name),
        schema_name=second_sub_schema_worksheet_name,
        metadata=[
            MetadataParams(key="Description", value=question_for_second_subschema),
            MetadataParams(
                key="Schema Type",
                value="Sub-Schema",
                value_restrictions=[
                    "Verifiable Credentials",
                    "Encrypted Verifiable Credential",
                    "Sub-Schema",
                ],
            ),
        ],
        table_columns=[
            FieldColumnParams(value="Required Field", value_restrictions=["Yes", "No"], width=20.0),
            FieldColumnParams(value="Field Type", width=40.0),
            FieldColumnParams(value="Parameter", width=20.0),
            FieldColumnParams(value="Visibility", width=15.0),
            FieldColumnParams(value="Question", width=70.0),
            FieldColumnParams(
                value="Allow Multiple Answers", value_restrictions=["Yes", "No"], width=30.0
            ),
            FieldColumnParams(value="Answer", width=50.0),
            FieldColumnParams(value="Key", width=15.0),
        ],
        table_rows=[
            [
                "Yes",
                "String",
                "",
                "TRUE",
                "What is your job title?",
                "No",
                "Software Engineer",
                "job_title",
            ]
        ],
    )

    enum_params = [
        ExcelPolicyEnumBuilderParams(
            sheet_name="CountryEnumSheet",
            schema_name="Test Schema Name",
            field_name="What is your country?",
            options=["USA", "Canada", "UK", "Australia", "Other"],
        )
    ]

    worksheet_name = "Test Schema Name"
    main_policy_schema_params = ExcelPolicySchemaParams(
        worksheet_name=WorksheetNameParams(value=worksheet_name),
        schema_name=worksheet_name,
        metadata=[
            MetadataParams(key="Description", value="Test schema for generating Excel from JSON"),
            MetadataParams(
                key="Schema Type",
                value="Verifiable Credentials",
                value_restrictions=[
                    "Verifiable Credentials",
                    "Encrypted Verifiable Credential",
                    "Sub-Schema",
                ],
            ),
        ],
        table_columns=[
            FieldColumnParams(value="Required Field", value_restrictions=["Yes", "No"], width=20.0),
            FieldColumnParams(value="Field Type", width=40.0),
            FieldColumnParams(value="Parameter", width=20.0),
            FieldColumnParams(value="Visibility", width=15.0),
            FieldColumnParams(value="Question", width=70.0),
            FieldColumnParams(
                value="Allow Multiple Answers", value_restrictions=["Yes", "No"], width=30.0
            ),
            FieldColumnParams(value="Answer", width=50.0),
            FieldColumnParams(value="Key", width=15.0),
        ],
        table_rows=[
            ["Yes", "String", "", "TRUE", "What is your full name?", "No", "John Doe", "full_name"],
            [
                "Yes",
                "String",
                "",
                "TRUE",
                "What is your email address?",
                "No",
                "john.doe@example.com",
                "email_address",
            ],
            ExcelPolicySchemaBuilder.SubRows(
                heading_row=ExcelPolicySchemaBuilder.Row(
                    "Yes",
                    ExcelPolicySchemaBuilder.SheetRefCell(
                        title="Sub Schema 1",
                        sheet_name=sub_schema_worksheet_name,
                    ),
                    "",
                    "TRUE",
                    question_for_subschema,
                    "No",
                    "Sub John Doe",
                    "sub_full_name",
                ),
                subrows=sub_policy_schema_params.table_rows.copy(),
            ),
            [
                "No",
                "Enum",
                ExcelPolicySchemaBuilder.SheetRefCell(
                    title="CountryEnum",
                    sheet_name="CountryEnumSheet",
                ),
                "TRUE",
                "What is your country?",
                "No",
                ExcelPolicySchemaBuilder.CellBase(enum_params[0].options[0]).set_value_restrictions(
                    enum_params[0].options
                ),
                "country",
            ],
        ],
    )

    builder = CompoundExcelBuilder(
        CompoundExcelBuilderParams(
            excel_policy_schema_params=[
                sub_policy_schema_params,  # Sub-schema first
                main_policy_schema_params,  # Root schema second
                second_sub_policy_schema_params,  # Another sub-schema last
            ],
            enum_params=enum_params,
            worksheet_file_name="compound_policy_schema_test.xlsx",
        )
    )
    builder.build()
    builder.save()
    print("Excel file created successfully (reverse order test)!")
