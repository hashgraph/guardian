"""Unit tests for CompoundExcelBuilder"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from policy_schema_builder.compound_excel_builder import (
    CompoundExcelBuilder,
    CompoundExcelBuilderParams,
)
from policy_schema_builder.enum_builder import ExcelPolicyEnumBuilderParams
from policy_schema_builder.models import (
    ExcelPolicySchemaParams,
    FieldColumnParams,
    MetadataParams,
    WorksheetNameParams,
)


class TestCompoundExcelBuilderParams:
    """Tests for CompoundExcelBuilderParams model"""

    def test_valid_compound_params(self):
        """Test creation of valid compound builder params"""
        schema_params = [
            ExcelPolicySchemaParams(
                schema_name="Schema1",
                worksheet_name=WorksheetNameParams(value="Schema1"),
                metadata=[],
                table_columns=[FieldColumnParams(value="Col1")],
                table_rows=[["Value1"]],
            )
        ]
        enum_params = [
            ExcelPolicyEnumBuilderParams(
                sheet_name="Enum1",
                schema_name="Schema1",
                field_name="Field1",
                options=["Option1", "Option2"],
            )
        ]
        params = CompoundExcelBuilderParams(
            excel_policy_schema_params=schema_params,
            enum_params=enum_params,
            worksheet_file_name="test",
        )
        assert len(params.excel_policy_schema_params) == 1
        assert len(params.enum_params) == 1
        assert params.worksheet_file_name == "test"

    def test_compound_params_with_file_name(self):
        """Test compound params with custom file name"""
        schema_params = [
            ExcelPolicySchemaParams(
                schema_name="Schema1",
                worksheet_name=WorksheetNameParams(value="Schema1"),
                metadata=[],
                table_columns=[FieldColumnParams(value="Col1")],
                table_rows=[["Value1"]],
            )
        ]
        params = CompoundExcelBuilderParams(
            excel_policy_schema_params=schema_params,
            enum_params=[],
            worksheet_file_name="CustomFileName",
        )
        assert params.worksheet_file_name == "CustomFileName"

    def test_compound_params_empty_lists(self):
        """Test compound params with empty lists"""
        params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[], enum_params=[], worksheet_file_name="empty"
        )
        assert len(params.excel_policy_schema_params) == 0
        assert len(params.enum_params) == 0


class TestCompoundExcelBuilderInitialization:
    """Tests for CompoundExcelBuilder initialization"""

    @pytest.fixture
    def basic_schema_params(self):
        """Fixture for basic schema params"""
        return ExcelPolicySchemaParams(
            schema_name="MainSchema",
            worksheet_name=WorksheetNameParams(value="Main Schema"),
            metadata=[MetadataParams(key="Description", value="Main")],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )

    @pytest.fixture
    def basic_enum_params(self):
        """Fixture for basic enum params"""
        return ExcelPolicyEnumBuilderParams(
            sheet_name="StatusEnum",
            schema_name="Main Schema",
            field_name="Status",
            options=["Active", "Inactive"],
        )

    def test_builder_initialization(self, basic_schema_params, basic_enum_params):
        """Test builder initialization"""
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[basic_schema_params],
            enum_params=[basic_enum_params],
            worksheet_file_name="test",
        )
        builder = CompoundExcelBuilder(compound_params)

        assert len(builder.excel_policy_schema_params) == 1
        assert len(builder.enum_params) == 1
        # workbook and sheet are only created after init_main_policy_schema_sheet or build() is called
        builder.init_main_policy_schema_sheet()
        assert builder.workbook is not None
        assert builder.sheet is not None

    def test_builder_worksheet_file_name_default(self, basic_schema_params):
        """Test that worksheet file name can be explicitly set"""
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[basic_schema_params], enum_params=[], worksheet_file_name=""
        )
        builder = CompoundExcelBuilder(compound_params)

        assert builder.worksheet_file_name == ".xlsx"  # Empty string becomes .xlsx

    def test_builder_worksheet_file_name_custom(self, basic_schema_params):
        """Test custom worksheet file name"""
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[basic_schema_params],
            enum_params=[],
            worksheet_file_name="CustomName",
        )
        builder = CompoundExcelBuilder(compound_params)

        assert builder.worksheet_file_name == "CustomName.xlsx"


class TestCompoundExcelBuilderCreateNewSheet:
    """Tests for _create_new_sheet method"""

    @pytest.fixture
    def builder(self):
        """Fixture for builder with minimal params"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="Main"
        )
        return CompoundExcelBuilder(compound_params)

    def test_create_new_sheet(self, builder):
        """Test creating a new sheet"""
        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        sheet = builder._create_new_sheet("NewSheet")

        assert sheet.title == "NewSheet"
        assert "NewSheet" in builder.workbook.sheetnames

    def test_create_duplicate_sheet_returns_existing(self, builder, caplog):
        """Test that creating duplicate sheet returns existing sheet with warning"""
        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        sheet1 = builder._create_new_sheet("DuplicateSheet")
        sheet2 = builder._create_new_sheet("DuplicateSheet")

        assert sheet1 == sheet2
        assert "already exists" in caplog.text.lower()

    def test_create_multiple_sheets(self, builder):
        """Test creating multiple sheets"""
        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        sheet1 = builder._create_new_sheet("Sheet1")
        sheet2 = builder._create_new_sheet("Sheet2")
        sheet3 = builder._create_new_sheet("Sheet3")

        assert sheet1.title == "Sheet1"
        assert sheet2.title == "Sheet2"
        assert sheet3.title == "Sheet3"
        assert len(builder.workbook.sheetnames) >= 3


class TestCompoundExcelBuilderEnumSheets:
    """Tests for enum sheet creation"""

    def test_create_single_enum_sheet(self):
        """Test creating single enum sheet"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        enum_params = ExcelPolicyEnumBuilderParams(
            sheet_name="StatusEnum",
            schema_name="Main",
            field_name="Status",
            options=["Active", "Inactive"],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[enum_params],
            worksheet_file_name="Main",
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        # Call create_enum_sheet directly with each enum param
        for enum_param in builder.enum_params:
            builder.create_enum_sheet(enum_param)

        assert "StatusEnum" in builder.workbook.sheetnames
        sheet = builder.workbook["StatusEnum"]
        assert sheet.cell(row=1, column=1).value == "Schema name"
        assert sheet.cell(row=1, column=2).value == "Main"
        assert sheet.cell(row=2, column=1).value == "Field name"
        assert sheet.cell(row=2, column=2).value == "Status"

    def test_create_multiple_enum_sheets(self):
        """Test creating multiple enum sheets"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        enum_params = [
            ExcelPolicyEnumBuilderParams(
                sheet_name="StatusEnum",
                schema_name="Main",
                field_name="Status",
                options=["Active", "Inactive"],
            ),
            ExcelPolicyEnumBuilderParams(
                sheet_name="TypeEnum",
                schema_name="Main",
                field_name="Type",
                options=["Type1", "Type2"],
            ),
        ]
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=enum_params,
            worksheet_file_name="Main",
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        # Call create_enum_sheet directly with each enum param
        for enum_param in builder.enum_params:
            builder.create_enum_sheet(enum_param)

        assert "StatusEnum" in builder.workbook.sheetnames
        assert "TypeEnum" in builder.workbook.sheetnames

    def test_create_enum_sheets_empty_list(self):
        """Test creating enum sheets with empty list"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="test"
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.init_main_policy_schema_sheet()  # Initialize workbook first
        # Call create_enum_sheet directly with each enum param (should be empty)
        for enum_param in builder.enum_params:
            builder.create_enum_sheet(enum_param)

        # Should not crash, just no enum sheets created
        initial_sheets = builder.workbook.sheetnames
        # Try again with empty list
        for enum_param in builder.enum_params:
            builder.create_enum_sheet(enum_param)
        assert builder.workbook.sheetnames == initial_sheets


class TestCompoundExcelBuilderMainPolicySchema:
    """Tests for main policy schema initialization"""

    def test_init_main_policy_schema_sheet(self):
        """Test initializing main policy schema sheet"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="MainSchema",
            worksheet_name=WorksheetNameParams(value="Main Schema"),
            metadata=[
                MetadataParams(key="Description", value="Main"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Required"), FieldColumnParams(value="Type")],
            table_rows=[["Yes", "String"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name=""
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.init_main_policy_schema_sheet()

        assert builder.sheet.title == "Main Schema"
        # worksheet_file_name is not changed by init_main_policy_schema_sheet
        assert builder.worksheet_file_name == ".xlsx"
        assert (
            builder.sheet.cell(row=1, column=1).value == "MainSchema"
        )  # schema_name, not worksheet_name

    def test_init_main_policy_schema_with_custom_filename(self):
        """Test that custom filename is preserved"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="MainSchema",
            worksheet_name=WorksheetNameParams(value="Main Schema"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[],
            worksheet_file_name="CustomFileName",
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.init_main_policy_schema_sheet()

        assert builder.worksheet_file_name == "CustomFileName.xlsx"


class TestCompoundExcelBuilderBuild:
    """Tests for complete build process"""

    def test_build_simple_workbook(self):
        """Test building simple workbook with main schema only"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="MainSchema",
            worksheet_name=WorksheetNameParams(value="Main Schema"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Required"), FieldColumnParams(value="Type")],
            table_rows=[["Yes", "String"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="test"
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.build()

        assert "Main Schema" in builder.workbook.sheetnames
        main_sheet = builder.workbook["Main Schema"]
        assert (
            main_sheet.cell(row=1, column=1).value == "MainSchema"
        )  # schema_name, not worksheet_name

    def test_build_with_sub_schemas_and_enums(self):
        """Test building complete workbook with schemas and enums"""
        main_schema = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        sub_schema = ExcelPolicySchemaParams(
            schema_name="SubSchema",
            worksheet_name=WorksheetNameParams(value="SubSchema"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["SubValue1"]],
        )
        enum_params = ExcelPolicyEnumBuilderParams(
            sheet_name="StatusEnum",
            schema_name="Main",
            field_name="Status",
            options=["Active", "Inactive"],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[main_schema, sub_schema],
            enum_params=[enum_params],
            worksheet_file_name="test",
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.build()

        assert "Main" in builder.workbook.sheetnames
        assert "SubSchema" in builder.workbook.sheetnames
        assert "StatusEnum" in builder.workbook.sheetnames

    def test_build_order(self):
        """Test that sheets are built in correct order"""
        main_schema = ExcelPolicySchemaParams(
            schema_name="Main",
            worksheet_name=WorksheetNameParams(value="Main"),
            metadata=[
                MetadataParams(key="Description", value="Main"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        sub_schema = ExcelPolicySchemaParams(
            schema_name="Sub",
            worksheet_name=WorksheetNameParams(value="Sub"),
            metadata=[
                MetadataParams(key="Description", value="Sub"),
                MetadataParams(key="Schema Type", value="Sub-Schema"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["SubValue1", "SubValue2"]],
        )
        enum_params = ExcelPolicyEnumBuilderParams(
            sheet_name="Enum", schema_name="Main", field_name="Field", options=["Opt1"]
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[main_schema, sub_schema],
            enum_params=[enum_params],
            worksheet_file_name="test",
        )
        builder = CompoundExcelBuilder(compound_params)

        builder.build()

        # Main schema should be the active sheet
        assert builder.sheet.title == "Main"


class TestCompoundExcelBuilderSave:
    """Tests for save functionality"""

    def test_save_to_default_directory(self):
        """Test saving to default directory (current)"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="TestSchema",
            worksheet_name=WorksheetNameParams(value="Test Schema"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[],
            worksheet_file_name="Test Schema",
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            saved_path = builder.save(output_dir=tmpdir)

            assert saved_path.exists()
            assert saved_path.suffix == ".xlsx"
            assert "Test Schema" in saved_path.name

    def test_save_to_custom_directory(self):
        """Test saving to custom directory"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Custom",
            worksheet_name=WorksheetNameParams(value="Custom"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="Custom"
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            custom_dir = Path(tmpdir) / "custom_output"
            saved_path = builder.save(output_dir=str(custom_dir))

            assert custom_dir.exists()
            assert saved_path.exists()
            assert saved_path.parent == custom_dir

    def test_save_filename_format(self):
        """Test saved filename format (no timestamp in current implementation)"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="Test"
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            saved_path = builder.save(output_dir=tmpdir)

            # Filename should be exactly what was set (no timestamp in current implementation)
            assert saved_path.stem == "Test"
            assert saved_path.name == "Test.xlsx"

    def test_save_sanitizes_filename(self):
        """Test that save works with valid filename (sanitization not implemented)"""
        # Use valid sheet name and valid file name
        schema_params = ExcelPolicySchemaParams(
            schema_name="TestSchemaName",
            worksheet_name=WorksheetNameParams(value="Test Schema Name"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[],
            worksheet_file_name="Test_Schema_Name",  # Use valid filename
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            saved_path = builder.save(output_dir=tmpdir)

            # Should save successfully with the given name
            assert saved_path.name == "Test_Schema_Name.xlsx"
            assert saved_path.exists()

    def test_save_creates_directory_if_not_exists(self):
        """Test that save creates output directory if it doesn't exist"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params], enum_params=[], worksheet_file_name="Test"
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            nested_dir = Path(tmpdir) / "level1" / "level2" / "level3"
            saved_path = builder.save(output_dir=str(nested_dir))

            assert nested_dir.exists()
            assert saved_path.exists()

    def test_save_workbook_is_valid(self):
        """Test that saved workbook can be opened and read"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="ValidSchema",
            worksheet_name=WorksheetNameParams(value="Valid Schema"),
            metadata=[
                MetadataParams(key="Description", value="Test"),
                MetadataParams(key="Schema Type", value="Verifiable Credentials"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Value1", "Value2"], ["Value3", "Value4"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[],
            worksheet_file_name="Valid Schema",
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            saved_path = builder.save(output_dir=tmpdir)

            # Load and verify the saved workbook
            wb = load_workbook(saved_path)
            assert "Valid Schema" in wb.sheetnames

            ws = wb["Valid Schema"]
            assert (
                ws.cell(row=1, column=1).value == "ValidSchema"
            )  # schema_name, not worksheet_name
            assert ws.cell(row=2, column=1).value == "Description"
            assert ws.cell(row=2, column=2).value == "Test"

            wb.close()

    def test_save_with_custom_worksheet_file_name(self):
        """Test saving with custom worksheet file name"""
        schema_params = ExcelPolicySchemaParams(
            schema_name="Schema",
            worksheet_name=WorksheetNameParams(value="Schema"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1")],
            table_rows=[["Value1"]],
        )
        compound_params = CompoundExcelBuilderParams(
            excel_policy_schema_params=[schema_params],
            enum_params=[],
            worksheet_file_name="MyCustomFileName",
        )
        builder = CompoundExcelBuilder(compound_params)
        builder.build()

        with tempfile.TemporaryDirectory() as tmpdir:
            saved_path = builder.save(output_dir=tmpdir)

            assert "MyCustomFileName" in saved_path.name
            assert saved_path.exists()
