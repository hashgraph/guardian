"""Unit tests for ExcelPolicySchemaBuilder"""

import pytest
from openpyxl import Workbook

from policy_schema_builder.models import (
    ExcelPolicySchemaParams,
    FieldColumnParams,
    MetadataParams,
    WorksheetNameParams,
)
from policy_schema_builder.policy_schema_builder import ExcelPolicySchemaBuilder


class TestExcelPolicySchemaBuilderRow:
    """Tests for ExcelPolicySchemaBuilder.Row class"""

    def test_row_creation_with_values(self):
        """Test creating a Row with values"""
        row = ExcelPolicySchemaBuilder.Row("Value1", "Value2", "Value3")
        assert len(row.values) == 3
        assert row.values == ["Value1", "Value2", "Value3"]

    def test_row_creation_empty(self):
        """Test creating an empty Row"""
        row = ExcelPolicySchemaBuilder.Row()
        assert len(row.values) == 0

    def test_row_to_list(self):
        """Test converting Row to list"""
        row = ExcelPolicySchemaBuilder.Row("A", "B", "C")
        result = row.to_list()
        assert result == ["A", "B", "C"]

    def test_row_add_cell(self):
        """Test adding cells to Row"""
        row = ExcelPolicySchemaBuilder.Row("A", "B")
        row.add_cell("C")
        assert len(row.values) == 3
        assert row.values[-1] == "C"

    def test_row_with_callable(self):
        """Test Row with callable value"""

        def custom_handler(cell):
            return "Custom"

        row = ExcelPolicySchemaBuilder.Row("A", custom_handler, "C")
        assert len(row.values) == 3
        assert callable(row.values[1])


class TestExcelPolicySchemaBuilderSubRows:
    """Tests for ExcelPolicySchemaBuilder.SubRows class"""

    def test_subrows_creation(self):
        """Test creating SubRows"""
        heading = ExcelPolicySchemaBuilder.Row("Parent1", "Parent2")
        subrows = [
            ExcelPolicySchemaBuilder.Row("Child1", "Child2"),
            ExcelPolicySchemaBuilder.Row("Child3", "Child4"),
        ]
        sub = ExcelPolicySchemaBuilder.SubRows(heading_row=heading, subrows=subrows)

        assert sub.heading_row == heading
        assert len(sub.subrows) == 2
        assert sub.is_nested is False

    def test_subrows_nested(self):
        """Test creating nested SubRows"""
        heading = ExcelPolicySchemaBuilder.Row("Parent1", "Parent2")
        subrows = [ExcelPolicySchemaBuilder.Row("Child1", "Child2")]
        sub = ExcelPolicySchemaBuilder.SubRows(heading_row=heading, subrows=subrows, is_nested=True)

        assert sub.is_nested is True

    def test_subrows_to_nested_list_simple(self):
        """Test converting SubRows to nested list"""
        heading = ExcelPolicySchemaBuilder.Row("A", "B")
        subrows = [ExcelPolicySchemaBuilder.Row("C", "D"), ExcelPolicySchemaBuilder.Row("E", "F")]
        sub = ExcelPolicySchemaBuilder.SubRows(heading_row=heading, subrows=subrows)

        result = sub.to_nested_list()

        assert len(result) == 3  # heading + 2 subrows
        assert result[0] == ["A", "B"]
        assert result[1] == ["C", "D"]
        assert result[2] == ["E", "F"]

    def test_subrows_to_nested_list_with_nested_flag(self):
        """Test converting nested SubRows with modified heading"""
        heading = ExcelPolicySchemaBuilder.Row("A", "B")
        subrows = [ExcelPolicySchemaBuilder.Row("C", "D")]
        sub = ExcelPolicySchemaBuilder.SubRows(heading_row=heading, subrows=subrows, is_nested=True)

        result = sub.to_nested_list()

        # When is_nested=True, heading cells are included as-is (no wrapping)
        assert len(result) == 2
        assert result[0] == ["A", "B"]  # Heading row values
        assert result[1] == ["C", "D"]  # Subrow values

    def test_subrows_deeply_nested(self):
        """Test deeply nested SubRows structure"""
        leaf_row = ExcelPolicySchemaBuilder.Row("Leaf1", "Leaf2")
        middle_subrows = ExcelPolicySchemaBuilder.SubRows(
            heading_row=ExcelPolicySchemaBuilder.Row("Middle1", "Middle2"), subrows=[leaf_row]
        )
        top_subrows = ExcelPolicySchemaBuilder.SubRows(
            heading_row=ExcelPolicySchemaBuilder.Row("Top1", "Top2"), subrows=[middle_subrows]
        )

        result = top_subrows.to_nested_list()

        assert len(result) == 2  # top heading + middle_subrows
        assert result[0] == ["Top1", "Top2"]
        assert isinstance(result[1], list)  # middle_subrows converted


class TestExcelPolicySchemaBuilderSheetRefCell:
    """Tests for ExcelPolicySchemaBuilder.SheetRefCell class"""

    def test_sheet_ref_cell_creation(self):
        """Test creating SheetRefCell"""
        ref = ExcelPolicySchemaBuilder.SheetRefCell(title="MyEnum", sheet_name="EnumSheet")
        assert ref.title == "MyEnum"
        assert ref.sheet_name == "EnumSheet"
        assert ref.focus_row == 1

    def test_sheet_ref_cell_with_focus_row(self):
        """Test SheetRefCell with custom focus row"""
        ref = ExcelPolicySchemaBuilder.SheetRefCell(
            title="MyEnum", sheet_name="EnumSheet", focus_row=5
        )
        assert ref.focus_row == 5

    def test_sheet_ref_cell_call(self):
        """Test calling SheetRefCell on a cell"""
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)

        ref = ExcelPolicySchemaBuilder.SheetRefCell(
            title="MyEnum", sheet_name="EnumSheet", focus_row=3
        )
        ref(cell, ws)

        assert cell.value == "MyEnum"
        assert cell.hyperlink is not None
        assert cell.hyperlink.location == "'EnumSheet'!A3"
        assert cell.font.underline == "single"


class TestExcelPolicySchemaBuilderBasics:
    """Tests for basic ExcelPolicySchemaBuilder functionality"""

    @pytest.fixture
    def workbook(self):
        """Fixture to create a fresh workbook"""
        return Workbook()

    @pytest.fixture
    def basic_params(self):
        """Fixture for basic schema params"""
        return ExcelPolicySchemaParams(
            schema_name="TestSchema",
            worksheet_name=WorksheetNameParams(value="Test Schema"),
            metadata=[MetadataParams(key="Description", value="Test description")],
            table_columns=[
                FieldColumnParams(value="Column1", width=20.0),
                FieldColumnParams(value="Column2", width=30.0),
            ],
            table_rows=[["Value1", "Value2"]],
        )

    def test_builder_initialization(self, workbook, basic_params):
        """Test builder initialization"""
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=basic_params, sheet=sheet)

        assert builder._params == basic_params
        assert builder._sheet == sheet
        assert builder._num_columns == 2

    def test_get_column_letter(self):
        """Test column index to letter conversion"""
        assert ExcelPolicySchemaBuilder.get_column_letter(1) == "A"
        assert ExcelPolicySchemaBuilder.get_column_letter(2) == "B"
        assert ExcelPolicySchemaBuilder.get_column_letter(26) == "Z"
        assert ExcelPolicySchemaBuilder.get_column_letter(27) == "AA"
        assert ExcelPolicySchemaBuilder.get_column_letter(52) == "AZ"

    def test_setup_column_widths(self, workbook, basic_params):
        """Test column width setup"""
        sheet = workbook.active
        _ = ExcelPolicySchemaBuilder(params=basic_params, sheet=sheet)

        assert sheet.column_dimensions["A"].width == 20.0
        assert sheet.column_dimensions["B"].width == 30.0

    def test_set_row_height(self, workbook, basic_params):
        """Test setting row height"""
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=basic_params, sheet=sheet)

        builder._set_row_height(1, 25.0)
        assert sheet.row_dimensions[1].height == 25.0

        builder._set_row_height(2)  # Default height
        assert sheet.row_dimensions[2].height == ExcelPolicySchemaBuilder.STANDARD_ROW_HEIGHT


class TestExcelPolicySchemaBuilderWorksheetName:
    """Tests for worksheet name functionality"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_set_worksheet_name(self, workbook):
        """Test setting worksheet name and title"""
        params = ExcelPolicySchemaParams(
            schema_name="MyTestSchema",
            worksheet_name=WorksheetNameParams(value="My Test Schema"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_worksheet_name(start_row=1)

        assert next_row == 2
        assert (
            sheet.cell(row=1, column=1).value == "MyTestSchema"
        )  # schema_name, not worksheet_name
        assert sheet.cell(row=1, column=1).font.bold is True
        assert sheet.cell(row=1, column=1).font.size == 14.0

    def test_worksheet_name_merged_cells(self, workbook):
        """Test that worksheet name is merged across all columns"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[
                FieldColumnParams(value="Col1"),
                FieldColumnParams(value="Col2"),
                FieldColumnParams(value="Col3"),
            ],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder._set_worksheet_name(start_row=1)

        # Check that cells are merged
        assert "A1:C1" in [str(merged) for merged in sheet.merged_cells.ranges]


class TestExcelPolicySchemaBuilderMetadata:
    """Tests for metadata functionality"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_set_metadata_simple(self, workbook):
        """Test setting simple metadata"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[
                MetadataParams(key="Description", value="Test description"),
                MetadataParams(key="Version", value="1.0"),
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_metadata(start_row=2)

        assert next_row == 4  # 2 metadata rows
        assert sheet.cell(row=2, column=1).value == "Description"
        assert sheet.cell(row=2, column=2).value == "Test description"
        assert sheet.cell(row=3, column=1).value == "Version"
        assert sheet.cell(row=3, column=2).value == "1.0"

    def test_set_metadata_with_restrictions(self, workbook):
        """Test metadata with value restrictions creates data validation"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[
                MetadataParams(
                    key="Schema Type", value="Type1", value_restrictions=["Type1", "Type2", "Type3"]
                )
            ],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder._set_metadata(start_row=2)

        # Check that data validation was added
        assert len(sheet.data_validations.dataValidation) > 0

    def test_metadata_merged_cells(self, workbook):
        """Test that metadata values are merged across columns"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[MetadataParams(key="Description", value="Test")],
            table_columns=[
                FieldColumnParams(value="Col1"),
                FieldColumnParams(value="Col2"),
                FieldColumnParams(value="Col3"),
            ],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder._set_metadata(start_row=2)

        # Check that value cells are merged
        assert "B2:C2" in [str(merged) for merged in sheet.merged_cells.ranges]


class TestExcelPolicySchemaBuilderTableHeaders:
    """Tests for table header functionality"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_set_fields_headers(self, workbook):
        """Test setting table column headers"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[
                FieldColumnParams(value="Required"),
                FieldColumnParams(value="Field Type"),
                FieldColumnParams(value="Question"),
            ],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_fields_headers(start_row=5)

        assert next_row == 6
        assert sheet.cell(row=5, column=1).value == "Required"
        assert sheet.cell(row=5, column=2).value == "Field Type"
        assert sheet.cell(row=5, column=3).value == "Question"

        # Check styling
        for col in range(1, 4):
            cell = sheet.cell(row=5, column=col)
            assert cell.font.bold is True
            assert cell.fill.start_color.rgb == ExcelPolicySchemaBuilder.TABLE_HEADER_FILL_COLOR


class TestExcelPolicySchemaBuilderDataRows:
    """Tests for data row functionality"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_set_fields_values_simple(self, workbook):
        """Test setting simple data rows"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[["Row1Col1", "Row1Col2"], ["Row2Col1", "Row2Col2"]],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_fields_values(start_row=1)

        assert next_row == 3  # 2 data rows
        assert sheet.cell(row=1, column=1).value == "Row1Col1"
        assert sheet.cell(row=1, column=2).value == "Row1Col2"
        assert sheet.cell(row=2, column=1).value == "Row2Col1"
        assert sheet.cell(row=2, column=2).value == "Row2Col2"

    def test_set_fields_values_with_row_objects(self, workbook):
        """Test setting data rows using Row objects"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[
                ExcelPolicySchemaBuilder.Row("A", "B"),
                ExcelPolicySchemaBuilder.Row("C", "D"),
            ],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_fields_values(start_row=1)

        assert next_row == 3
        assert sheet.cell(row=1, column=1).value == "A"
        assert sheet.cell(row=1, column=2).value == "B"

    def test_set_fields_values_with_callable(self, workbook):
        """Test data rows with callable cell values"""

        def custom_handler(cell, worksheet):
            cell.value = "Custom Value"

        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[[custom_handler, "Normal Value"]],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder._set_fields_values(start_row=1)

        assert sheet.cell(row=1, column=1).value == "Custom Value"
        assert sheet.cell(row=1, column=2).value == "Normal Value"

    def test_set_fields_values_with_nested_rows(self, workbook):
        """Test nested rows for outline grouping"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[
                ExcelPolicySchemaBuilder.Row("Parent1", "Parent2"),
                ExcelPolicySchemaBuilder.SubRows(
                    heading_row=ExcelPolicySchemaBuilder.Row("SubParent1", "SubParent2"),
                    subrows=[ExcelPolicySchemaBuilder.Row("Child1", "Child2")],
                ),
            ],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        next_row = builder._set_fields_values(start_row=1)

        assert next_row == 4  # Parent + SubParent heading + Child
        assert sheet.cell(row=1, column=1).value == "Parent1"
        assert sheet.cell(row=2, column=1).value == "SubParent1"
        assert sheet.cell(row=3, column=1).value == "Child1"

        # Check outline levels
        assert sheet.row_dimensions[1].outline_level == 0
        assert sheet.row_dimensions[2].outline_level == 0
        assert sheet.row_dimensions[3].outline_level == 1

    def test_is_nested_row(self, workbook):
        """Test nested row detection"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Col1"), FieldColumnParams(value="Col2")],
            table_rows=[],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        # Regular row
        assert builder._is_nested_row(["A", "B"]) is False

        # Nested row
        assert builder._is_nested_row([["A", "B"], ["C", "D"]]) is True

        # Not a list
        assert builder._is_nested_row("not a list") is False


class TestExcelPolicySchemaBuilderValidation:
    """Tests for data validation functionality"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_set_fields_validations(self, workbook):
        """Test setting data validations on columns"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[
                FieldColumnParams(value="Required", value_restrictions=["Yes", "No"]),
                FieldColumnParams(value="Field Type"),
            ],
            table_rows=[["Yes", "String"], ["No", "Integer"]],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder._set_fields_validations(start_row=1, end_row=2)

        # Check that data validation was added
        assert len(sheet.data_validations.dataValidation) > 0


class TestExcelPolicySchemaBuilderBuild:
    """Tests for complete build process"""

    @pytest.fixture
    def workbook(self):
        return Workbook()

    def test_build_complete_schema(self, workbook):
        """Test building complete schema"""
        params = ExcelPolicySchemaParams(
            schema_name="CompleteSchema",
            worksheet_name=WorksheetNameParams(value="Complete Schema"),
            metadata=[
                MetadataParams(key="Description", value="Complete test"),
                MetadataParams(key="Version", value="1.0"),
            ],
            table_columns=[
                FieldColumnParams(value="Required", width=15.0),
                FieldColumnParams(value="Type", width=20.0),
                FieldColumnParams(value="Question", width=50.0),
            ],
            table_rows=[
                ["Yes", "String", "What is your name?"],
                ["No", "Integer", "What is your age?"],
            ],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        result = builder.build()

        assert result == builder

        # Verify worksheet name (displays schema_name)
        assert sheet.cell(row=1, column=1).value == "CompleteSchema"

        # Verify metadata
        assert sheet.cell(row=2, column=1).value == "Description"
        assert sheet.cell(row=3, column=1).value == "Version"

        # Verify headers (after name + 2 metadata rows)
        assert sheet.cell(row=4, column=1).value == "Required"
        assert sheet.cell(row=4, column=2).value == "Type"
        assert sheet.cell(row=4, column=3).value == "Question"

        # Verify data
        assert sheet.cell(row=5, column=1).value == "Yes"
        assert sheet.cell(row=5, column=3).value == "What is your name?"
        assert sheet.cell(row=6, column=1).value == "No"
        assert sheet.cell(row=6, column=3).value == "What is your age?"

    def test_build_with_hyperlinks(self, workbook):
        """Test building with SheetRefCell hyperlinks"""
        params = ExcelPolicySchemaParams(
            schema_name="Test",
            worksheet_name=WorksheetNameParams(value="Test"),
            metadata=[],
            table_columns=[FieldColumnParams(value="Type"), FieldColumnParams(value="Parameter")],
            table_rows=[
                ExcelPolicySchemaBuilder.Row(
                    "Enum",
                    ExcelPolicySchemaBuilder.SheetRefCell(
                        title="StatusEnum", sheet_name="StatusEnumSheet"
                    ),
                )
            ],
        )
        sheet = workbook.active
        builder = ExcelPolicySchemaBuilder(params=params, sheet=sheet)

        builder.build()

        # Check hyperlink was created
        # Row 1: Title, Row 2: Table header, Row 3: Data row
        cell = sheet.cell(row=3, column=2)
        assert cell.value == "StatusEnum"
        assert cell.hyperlink is not None
        assert "'StatusEnumSheet'!A1" in cell.hyperlink.location
