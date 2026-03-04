"""Unit tests for ExcelPolicyEnumBuilder"""

import pytest
from openpyxl import Workbook

from policy_schema_builder.enum_builder import (
    ExcelPolicyEnumBuilder,
    ExcelPolicyEnumBuilderParams,
)


class TestExcelPolicyEnumBuilderParams:
    """Tests for ExcelPolicyEnumBuilderParams model"""

    def test_valid_enum_params(self):
        """Test creation of valid enum builder params"""
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="TestEnum",
            schema_name="Test Schema",
            field_name="Test Field",
            options=["Option1", "Option2", "Option3"],
        )
        assert params.sheet_name == "TestEnum"
        assert params.schema_name == "Test Schema"
        assert params.field_name == "Test Field"
        assert len(params.options) == 3

    def test_enum_params_with_empty_options(self):
        """Test enum params with empty options list"""
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="TestEnum", schema_name="Test Schema", field_name="Test Field", options=[]
        )
        assert len(params.options) == 0

    def test_enum_params_required_fields(self):
        """Test that all required fields are enforced"""
        from pydantic import ValidationError

        with pytest.raises(ValidationError):
            ExcelPolicyEnumBuilderParams(
                sheet_name="TestEnum"
                # Missing schema_name, field_name, and options
            )


class TestExcelPolicyEnumBuilder:
    """Tests for ExcelPolicyEnumBuilder class"""

    @pytest.fixture
    def workbook(self):
        """Fixture to create a fresh workbook for each test"""
        return Workbook()

    @pytest.fixture
    def basic_enum_params(self):
        """Fixture for basic enum parameters"""
        return ExcelPolicyEnumBuilderParams(
            sheet_name="StatusEnum",
            schema_name="Employee Schema",
            field_name="Employment Status",
            options=["Full-time", "Part-time", "Contractor"],
        )

    def test_builder_initialization(self, workbook, basic_enum_params):
        """Test builder initialization"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        assert builder.sheet == sheet
        assert builder.params == basic_enum_params
        assert builder.options_start_row == 0

    def test_setup_column_widths(self, workbook, basic_enum_params):
        """Test column width setup"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)
        builder._setup_column_widths()

        assert sheet.column_dimensions["A"].width == 30.0
        assert sheet.column_dimensions["B"].width == 50.0

    def test_apply_header_title_styling(self, workbook, basic_enum_params):
        """Test header title cell styling"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        cell = sheet.cell(row=1, column=1, value="Test Header")
        builder._apply_header_title(cell)

        assert cell.font.bold is True
        assert cell.font.size == 14
        assert cell.font.name == "Calibri"
        assert cell.border.left is not None
        assert cell.border.right is not None
        assert cell.border.top is not None
        assert cell.border.bottom is not None

    def test_apply_value_cell_styling(self, workbook, basic_enum_params):
        """Test value cell styling"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        cell = sheet.cell(row=1, column=1, value="Test Value")
        builder._apply_value_cell(cell)

        assert cell.font.bold is False
        assert cell.font.size == 11
        assert cell.font.name == "Calibri"
        assert cell.alignment.wrap_text is True
        assert cell.border.left is not None
        assert cell.border.right is not None
        assert cell.border.top is not None
        assert cell.border.bottom is not None

    def test_apply_option_cell_styling(self, workbook, basic_enum_params):
        """Test option cell styling"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        cell = sheet.cell(row=1, column=1, value="Option")
        builder._apply_option_cell(cell)

        assert cell.font.bold is False
        assert cell.font.size == 11
        assert cell.font.name == "Calibri"
        assert cell.alignment.wrap_text is True
        assert cell.border.left is not None
        assert cell.border.right is not None

    def test_set_schema_name(self, workbook, basic_enum_params):
        """Test schema name row creation"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        rows_used = builder._set_schema_name()

        assert rows_used == 1
        assert sheet.cell(row=1, column=1).value == "Schema name"
        assert sheet.cell(row=1, column=2).value == "Employee Schema"
        assert sheet.cell(row=1, column=1).font.bold is True

    def test_set_field_name(self, workbook, basic_enum_params):
        """Test field name row creation"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        rows_used = builder._set_field_name(start_row=2)

        assert rows_used == 1
        assert sheet.cell(row=2, column=1).value == "Field name"
        assert sheet.cell(row=2, column=2).value == "Employment Status"
        assert sheet.cell(row=2, column=1).font.bold is True

    def test_set_enum_options(self, workbook, basic_enum_params):
        """Test enum options creation with merging"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        rows_used = builder._set_enum_options(start_row=3)

        assert rows_used == 3  # Three options
        assert sheet.cell(row=3, column=1).value == "Full-time"
        assert sheet.cell(row=4, column=1).value == "Part-time"
        assert sheet.cell(row=5, column=1).value == "Contractor"

        # Check that cells are merged
        assert "A3:B3" in [str(merged) for merged in sheet.merged_cells.ranges]
        assert "A4:B4" in [str(merged) for merged in sheet.merged_cells.ranges]
        assert "A5:B5" in [str(merged) for merged in sheet.merged_cells.ranges]

    def test_set_enum_options_with_empty_list(self, workbook):
        """Test enum options with empty list"""
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="EmptyEnum", schema_name="Test Schema", field_name="Test Field", options=[]
        )
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=params)

        rows_used = builder._set_enum_options(start_row=3)

        assert rows_used == 0

    def test_build_complete_enum_sheet(self, workbook, basic_enum_params):
        """Test building complete enum sheet"""
        sheet = workbook.create_sheet(title="StatusEnum")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        result_sheet = builder.build()

        assert result_sheet == sheet
        assert builder.options_start_row == 3

        # Verify schema name row
        assert sheet.cell(row=1, column=1).value == "Schema name"
        assert sheet.cell(row=1, column=2).value == "Employee Schema"

        # Verify field name row
        assert sheet.cell(row=2, column=1).value == "Field name"
        assert sheet.cell(row=2, column=2).value == "Employment Status"

        # Verify options
        assert sheet.cell(row=3, column=1).value == "Full-time"
        assert sheet.cell(row=4, column=1).value == "Part-time"
        assert sheet.cell(row=5, column=1).value == "Contractor"

        # Verify column widths
        assert sheet.column_dimensions["A"].width == 30.0
        assert sheet.column_dimensions["B"].width == 50.0

    def test_get_options_start_position(self, workbook, basic_enum_params):
        """Test getting options start position"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)
        builder.build()

        position = builder.get_options_start_position()

        assert position == "A3"

    def test_build_with_many_options(self, workbook):
        """Test building enum with many options"""
        many_options = [f"Option{i}" for i in range(1, 21)]  # 20 options
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="ManyOptionsEnum",
            schema_name="Test Schema",
            field_name="Many Options Field",
            options=many_options,
        )
        sheet = workbook.create_sheet(title="ManyOptionsEnum")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=params)

        builder.build()

        # Verify all options are present
        for i, option in enumerate(many_options, start=3):
            assert sheet.cell(row=i, column=1).value == option

        # Verify merging for all options
        for i in range(3, 23):  # rows 3 to 22
            assert f"A{i}:B{i}" in [str(merged) for merged in sheet.merged_cells.ranges]

    def test_build_with_unicode_options(self, workbook):
        """Test building enum with unicode characters in options"""
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="UnicodeEnum",
            schema_name="Test Schema",
            field_name="Unicode Field",
            options=["Español", "Français", "日本語", "中文"],
        )
        sheet = workbook.create_sheet(title="UnicodeEnum")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=params)

        builder.build()

        assert sheet.cell(row=3, column=1).value == "Español"
        assert sheet.cell(row=4, column=1).value == "Français"
        assert sheet.cell(row=5, column=1).value == "日本語"
        assert sheet.cell(row=6, column=1).value == "中文"

    def test_build_with_long_option_names(self, workbook):
        """Test building enum with very long option names"""
        params = ExcelPolicyEnumBuilderParams(
            sheet_name="LongOptionsEnum",
            schema_name="Test Schema",
            field_name="Long Field",
            options=[
                "This is a very long option name that should wrap properly in the cell",
                "Another extremely long option name to test the wrapping functionality",
            ],
        )
        sheet = workbook.create_sheet(title="LongOptionsEnum")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=params)

        builder.build()

        # Verify wrap text is enabled
        assert sheet.cell(row=3, column=1).alignment.wrap_text is True
        assert sheet.cell(row=4, column=1).alignment.wrap_text is True

    def test_build_preserves_existing_sheet(self, workbook, basic_enum_params):
        """Test that building doesn't create new sheet but uses provided one"""
        sheet = workbook.create_sheet(title="ExistingSheet")
        initial_sheet_count = len(workbook.sheetnames)

        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)
        builder.build()

        assert len(workbook.sheetnames) == initial_sheet_count
        assert "ExistingSheet" in workbook.sheetnames

    def test_multiple_builds_on_same_sheet(self, workbook, basic_enum_params):
        """Test building twice on the same sheet overwrites content"""
        sheet = workbook.create_sheet(title="TestSheet")
        builder = ExcelPolicyEnumBuilder(sheet=sheet, params=basic_enum_params)

        # First build
        builder.build()
        first_value = sheet.cell(row=1, column=2).value

        # Second build with different params
        new_params = ExcelPolicyEnumBuilderParams(
            sheet_name="NewEnum",
            schema_name="New Schema",
            field_name="New Field",
            options=["New1", "New2"],
        )
        builder2 = ExcelPolicyEnumBuilder(sheet=sheet, params=new_params)
        builder2.build()

        # Verify second build overwrote the first
        assert sheet.cell(row=1, column=2).value == "New Schema"
        assert sheet.cell(row=1, column=2).value != first_value
