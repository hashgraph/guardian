#!/usr/bin/env python3
"""
Excel Artifact Extraction Tool for Guardian Methodology Digitization Handbook

This tool extracts data from Excel files in the artifacts folder to ensure accuracy
in handbook examples and references. It provides functionality to list workbooks,
extract tab information, and extract specific content for use in handbook chapters.

Usage:
    python excel_artifact_extractor.py list-workbooks
    python excel_artifact_extractor.py extract-tabs <workbook_name>
    python excel_artifact_extractor.py extract-tab-content <workbook_name> <tab_name>
    python excel_artifact_extractor.py extract-schema-structure <workbook_name>
    python excel_artifact_extractor.py extract-parameters <workbook_name> <tab_name>
"""

import argparse
import os
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from typing import Dict, List, Any, Optional

class ExcelArtifactExtractor:
    """
    Excel artifact extraction tool for Guardian Methodology Digitization Handbook.
    
    Provides comprehensive functionality to extract data from Excel artifacts
    to ensure accuracy in handbook content and examples.
    """
    
    def __init__(self, artifacts_folder_path: str = None):
        """
        Initialize the extractor with artifacts folder path.
        
        Args:
            artifacts_folder_path: Path to artifacts folder. If None, uses current directory.
        """
        if artifacts_folder_path is None:
            artifacts_folder_path = Path(__file__).parent
        self.artifacts_path = Path(artifacts_folder_path)
        
        if not self.artifacts_path.exists():
            raise FileNotFoundError(f"Artifacts folder not found: {self.artifacts_path}")
    
    def list_workbooks(self) -> List[str]:
        """
        List all Excel files in the artifacts folder.
        
        Returns:
            List of Excel workbook filenames
        """
        excel_extensions = {'.xlsx', '.xls', '.xlsm'}
        workbooks = []
        
        for file_path in self.artifacts_path.iterdir():
            if file_path.suffix.lower() in excel_extensions:
                workbooks.append(file_path.name)
        
        return sorted(workbooks)
    
    def extract_tabs(self, workbook_name: str) -> List[Dict[str, Any]]:
        """
        Extract all tab names and structure information from a workbook.
        
        Args:
            workbook_name: Name of the Excel workbook
            
        Returns:
            List of dictionaries containing tab information
        """
        workbook_path = self.artifacts_path / workbook_name
        
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            tabs_info = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get basic sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Try to determine content type based on sheet name and structure
                content_type = self._determine_content_type(sheet_name, sheet)
                
                # Get a sample of data to understand structure
                sample_data = []
                for row_idx in range(1, min(6, max_row + 1)):  # First 5 rows
                    row_data = []
                    for col_idx in range(1, min(11, max_col + 1)):  # First 10 columns
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        row_data.append(str(cell_value) if cell_value is not None else '')
                    sample_data.append(row_data)
                
                tab_info = {
                    'name': sheet_name,
                    'content_type': content_type,
                    'dimensions': {
                        'rows': max_row,
                        'columns': max_col
                    },
                    'sample_data': sample_data
                }
                tabs_info.append(tab_info)
            
            workbook.close()
            return tabs_info
            
        except Exception as e:
            raise Exception(f"Error extracting tabs from {workbook_name}: {str(e)}")
    
    def extract_tab_content(self, workbook_name: str, tab_name: str, 
                           max_rows: int = 100, max_cols: int = 50) -> Dict[str, Any]:
        """
        Extract content from a specific tab.
        
        Args:
            workbook_name: Name of the Excel workbook
            tab_name: Name of the worksheet tab
            max_rows: Maximum number of rows to extract
            max_cols: Maximum number of columns to extract
            
        Returns:
            Dictionary containing tab content and metadata
        """
        workbook_path = self.artifacts_path / workbook_name
        
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        try:
            # Use pandas for easier data manipulation
            df = pd.read_excel(workbook_path, sheet_name=tab_name, 
                             header=None, nrows=max_rows)
            
            # Limit columns
            if len(df.columns) > max_cols:
                df = df.iloc[:, :max_cols]
            
            # Convert to list of lists for JSON serialization
            content_data = df.fillna('').astype(str).values.tolist()
            
            # Also get column headers if they exist
            headers = []
            if len(content_data) > 0:
                headers = content_data[0]  # First row as potential headers
            
            return {
                'workbook': workbook_name,
                'tab_name': tab_name,
                'dimensions': {
                    'rows': len(content_data),
                    'columns': len(content_data[0]) if content_data else 0
                },
                'headers': headers,
                'content': content_data,
                'content_type': self._determine_content_type(tab_name, None)
            }
            
        except Exception as e:
            raise Exception(f"Error extracting content from {tab_name} in {workbook_name}: {str(e)}")
    
    def extract_schema_structure(self, workbook_name: str) -> Dict[str, Any]:
        """
        Extract Guardian schema structure from Excel workbook.
        
        Args:
            workbook_name: Name of the Excel workbook
            
        Returns:
            Dictionary containing schema structure information
        """
        workbook_path = self.artifacts_path / workbook_name
        
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        schema_info = {
            'workbook': workbook_name,
            'schemas': [],
            'extraction_metadata': {
                'tool': 'excel_artifact_extractor',
                'version': '1.0'
            }
        }
        
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                if self._is_schema_sheet(sheet_name):
                    sheet = workbook[sheet_name]
                    schema_structure = self._extract_schema_from_sheet(sheet, sheet_name)
                    if schema_structure:
                        schema_info['schemas'].append(schema_structure)
            
            workbook.close()
            return schema_info
            
        except Exception as e:
            raise Exception(f"Error extracting schema structure from {workbook_name}: {str(e)}")
    
    def extract_parameters(self, workbook_name: str, tab_name: str) -> Dict[str, Any]:
        """
        Extract parameter lists, data types, and validation rules from Excel tab.
        
        Args:
            workbook_name: Name of the Excel workbook
            tab_name: Name of the worksheet tab
            
        Returns:
            Dictionary containing parameter information
        """
        workbook_path = self.artifacts_path / workbook_name
        
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        try:
            df = pd.read_excel(workbook_path, sheet_name=tab_name)
            
            parameters = []
            
            # Look for common parameter column patterns
            param_columns = self._identify_parameter_columns(df)
            
            if param_columns:
                for index, row in df.iterrows():
                    param_info = {}
                    for col_type, col_name in param_columns.items():
                        if col_name in df.columns:
                            value = row[col_name]
                            if pd.notna(value):
                                param_info[col_type] = str(value)
                    
                    if param_info:  # Only add if we found parameter information
                        parameters.append(param_info)
            
            return {
                'workbook': workbook_name,
                'tab_name': tab_name,
                'parameter_columns': param_columns,
                'parameters': parameters,
                'total_parameters': len(parameters)
            }
            
        except Exception as e:
            raise Exception(f"Error extracting parameters from {tab_name} in {workbook_name}: {str(e)}")
    
    def _determine_content_type(self, sheet_name: str, sheet=None) -> str:
        """Determine the type of content in a sheet based on name and structure."""
        name_lower = sheet_name.lower()
        
        if any(keyword in name_lower for keyword in ['schema', 'pdd', 'mr', 'monitoring']):
            return 'guardian-schema'
        elif any(keyword in name_lower for keyword in ['test', 'calc', 'calculation', 'validation']):
            return 'validation-data'
        elif any(keyword in name_lower for keyword in ['param', 'input', 'output']):
            return 'parameter-data'
        elif any(keyword in name_lower for keyword in ['tool', 'ar-tool', 'cdm']):
            return 'tool-integration'
        else:
            return 'general-data'
    
    def _is_schema_sheet(self, sheet_name: str) -> bool:
        """Check if a sheet contains Guardian schema information."""
        name_lower = sheet_name.lower()
        return any(keyword in name_lower for keyword in ['schema', 'pdd', 'mr', 'monitoring'])
    
    def _extract_schema_from_sheet(self, sheet, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Extract Guardian schema structure from a worksheet."""
        try:
            # Basic schema extraction - this would need to be enhanced based on actual schema formats
            schema_structure = {
                'name': sheet_name,
                'type': self._determine_content_type(sheet_name),
                'fields': [],
                'dimensions': {
                    'rows': sheet.max_row,
                    'columns': sheet.max_column
                }
            }
            
            # Extract field information (assumes first row contains field names)
            if sheet.max_row > 0:
                for col_idx in range(1, sheet.max_column + 1):
                    field_name = sheet.cell(row=1, column=col_idx).value
                    if field_name:
                        schema_structure['fields'].append({
                            'name': str(field_name),
                            'column': col_idx
                        })
            
            return schema_structure if schema_structure['fields'] else None
            
        except Exception:
            return None
    
    def _identify_parameter_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """Identify columns that contain parameter information."""
        param_columns = {}
        
        # Common column name patterns
        column_patterns = {
            'name': ['parameter', 'param', 'name', 'variable', 'field'],
            'type': ['type', 'data_type', 'datatype', 'format'],
            'description': ['description', 'desc', 'comment', 'note'],
            'unit': ['unit', 'units', 'measurement'],
            'default': ['default', 'default_value', 'initial'],
            'validation': ['validation', 'constraint', 'rule']
        }
        
        for col_name in df.columns:
            col_name_lower = str(col_name).lower()
            for param_type, patterns in column_patterns.items():
                if any(pattern in col_name_lower for pattern in patterns):
                    param_columns[param_type] = col_name
                    break
        
        return param_columns


def main():
    """Command-line interface for the Excel Artifact Extractor."""
    parser = argparse.ArgumentParser(
        description='Excel Artifact Extraction Tool for Guardian Methodology Digitization Handbook',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python excel_artifact_extractor.py list-workbooks
    python excel_artifact_extractor.py extract-tabs VM0033_Allcot_Test_Case_Artifact.xlsx
    python excel_artifact_extractor.py extract-tab-content VM0033_Allcot_Test_Case_Artifact.xlsx "Test Data"
    python excel_artifact_extractor.py extract-schema-structure schema-template-excel.xlsx
    python excel_artifact_extractor.py extract-parameters VM0033_Allcot_Test_Case_Artifact.xlsx "Parameters"
        """
    )
    
    parser.add_argument('command', choices=[
        'list-workbooks', 'extract-tabs', 'extract-tab-content', 
        'extract-schema-structure', 'extract-parameters'
    ], help='Command to execute')
    
    parser.add_argument('workbook', nargs='?', help='Excel workbook filename')
    parser.add_argument('tab', nargs='?', help='Worksheet tab name')
    parser.add_argument('--artifacts-path', help='Path to artifacts folder')
    parser.add_argument('--output-format', choices=['json', 'pretty'], 
                       default='pretty', help='Output format')
    
    args = parser.parse_args()
    
    try:
        extractor = ExcelArtifactExtractor(args.artifacts_path)
        result = None
        
        if args.command == 'list-workbooks':
            result = extractor.list_workbooks()
            
        elif args.command == 'extract-tabs':
            if not args.workbook:
                parser.error('extract-tabs requires workbook argument')
            result = extractor.extract_tabs(args.workbook)
            
        elif args.command == 'extract-tab-content':
            if not args.workbook or not args.tab:
                parser.error('extract-tab-content requires workbook and tab arguments')
            result = extractor.extract_tab_content(args.workbook, args.tab)
            
        elif args.command == 'extract-schema-structure':
            if not args.workbook:
                parser.error('extract-schema-structure requires workbook argument')
            result = extractor.extract_schema_structure(args.workbook)
            
        elif args.command == 'extract-parameters':
            if not args.workbook or not args.tab:
                parser.error('extract-parameters requires workbook and tab arguments')
            result = extractor.extract_parameters(args.workbook, args.tab)
        
        # Output results
        if args.output_format == 'json':
            print(json.dumps(result, indent=2, ensure_ascii=False))
        else:
            # Pretty print format
            if args.command == 'list-workbooks':
                print("Available Excel Workbooks:")
                for workbook in result:
                    print(f"  - {workbook}")
                    
            elif args.command == 'extract-tabs':
                print(f"Tabs in {args.workbook}:")
                for tab in result:
                    print(f"  - {tab['name']} ({tab['content_type']}) - {tab['dimensions']['rows']}x{tab['dimensions']['columns']}")
                    
            elif args.command in ['extract-tab-content', 'extract-schema-structure', 'extract-parameters']:
                print(json.dumps(result, indent=2, ensure_ascii=False))
        
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()